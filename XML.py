import os
import sys
import shutil
import re
import threading
import time
import datetime
import socket
import base64
import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import webbrowser

def resource_path(relative_path):
    """Obtém o caminho correto para recursos, funcionando tanto em desenvolvimento quanto em executável"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def pluralizar(quantidade, singular, plural=None):
    """Retorna texto com pluralidade dinâmica baseada na quantidade"""
    if plural is None:
        plural = singular + "s"
    
    if quantidade == 1:
        return f"{quantidade} {singular}"
    else:
        return f"{quantidade} {plural}"

# --- IMPORTS PARA WEB AUTOMATION (UPLOAD) ---
try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.edge.options import Options
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
except ImportError:
    messagebox.showerror("Erro de Dependência",
                         "A biblioteca 'selenium' não está instalada.\nAbra o terminal e digite: pip install selenium")
    sys.exit()

# --- IMPORTS PARA LOGS EM EXCEL ---
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    messagebox.showerror("Erro de Dependência",
                         "A biblioteca 'openpyxl' não está instalada.\nAbra o terminal e digite: pip install openpyxl")
    sys.exit()


class AutoScrollbar(ttk.Scrollbar):
    """Barra de rolagem que se esconde se não for necessária."""

    def set(self, lo, hi):
        if float(lo) <= 0.0 and float(hi) >= 1.0:
            # Não há conteúdo suficiente para rolagem - esconder
            self.pack_forget()
        else:
            # Há conteúdo suficiente - mostrar scrollbar
            self.pack(side=tk.RIGHT, fill=tk.Y)
        ttk.Scrollbar.set(self, lo, hi)
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# ==============================================================================
# --- CONFIGURAÇÕES DINÂMICAS E PASTAS ---
# ==============================================================================
USER_HOME = os.path.expanduser("~")
CAMINHO_RELATIVO = r"DPDHL\Starlink Imp - Starlink_Imp"
BASE_DIR = os.path.join(USER_HOME, CAMINHO_RELATIVO)
URL_SHAREPOINT = "https://dpdhl.sharepoint.com/teams/StarlinkImp/Shared%20Documents"

URL_DHL_LINK = "https://link-cc.dhl.com/upload"

if not os.path.exists(BASE_DIR):
    def abrir_sharepoint(event): webbrowser.open(URL_SHAREPOINT)


    err_window = tk.Tk()
    err_window.title("Ação Necessária - Sincronização")
    largura, altura = 550, 750
    pos_x = (err_window.winfo_screenwidth() // 2) - (largura // 2)
    pos_y = (err_window.winfo_screenheight() // 2) - (altura // 2)
    err_window.geometry(f"{largura}x{altura}+{pos_x}+{pos_y}")
    err_window.configure(bg="white")

    try:
        err_window.iconbitmap(resource_path("logo.ico"))
    except:
        pass
    
    # Permitir fechar com ESC
    err_window.bind('<Escape>', lambda e: sys.exit())
    err_window.focus_set()

    tk.Label(err_window, text="PASTA NÃO SINCRONIZADA",
             fg="#D40511", bg="white", font=("Segoe UI", 16, "bold")).pack(pady=(15, 5))
    msg = (f"O sistema não encontrou a pasta 'Arquivos' no seu computador.\n"
           f"Caminho esperado: {BASE_DIR}")
    tk.Label(err_window, text=msg, bg="white", font=("Segoe UI", 10)).pack(pady=5)

    try:
        img_path = resource_path("tutorial_sync.png")
        png_tutorial = tk.PhotoImage(file=img_path)
        lbl_img = tk.Label(err_window, image=png_tutorial, bg="white", bd=1, relief="solid")
        lbl_img.image = png_tutorial
        lbl_img.pack(pady=10)
    except Exception:
        tk.Label(err_window, text="[Imagem tutorial_sync.png não encontrada]", bg="white", fg="gray").pack()

    instrucoes = ("Você deve clicar nos 3 pontos na parte superior do Sharepoint "
                  "e clicar no botão Sincronizar (ou Sync).\n\n"
                  "Um popup irá surgir pedindo permissão para sincronizar. "
                  "Após isso, o programa deve ser reiniciado.")
    tk.Label(err_window, text=instrucoes, bg="white", fg="#333",
             font=("Segoe UI", 10), justify="center", wraplength=500).pack(pady=5)

    lbl_link = tk.Label(err_window, text=">>> CLIQUE AQUI PARA ABRIR O SHAREPOINT <<<",
                        fg="blue", bg="white", cursor="hand2", font=("Segoe UI", 11, "bold", "underline"))
    lbl_link.pack(pady=15)
    lbl_link.bind("<Button-1>", abrir_sharepoint)

    tk.Button(err_window, text="Fechar Programa", command=sys.exit,
              bg="#333", fg="white", font=("Segoe UI", 10), width=20).pack(pady=10)
    err_window.mainloop()
    sys.exit()

DIR_ENTRADA = os.path.join(BASE_DIR, "1_XML_Origem")
DIR_DESTINO = os.path.join(BASE_DIR, "2_XML_Convertido")
DIR_ENVIADOS = os.path.join(BASE_DIR, "2.1_XML_Enviados")
DIR_LOGS = os.path.join(BASE_DIR, "3_Logs")

for pasta in [DIR_ENTRADA, DIR_DESTINO, DIR_ENVIADOS, DIR_LOGS]:
    if not os.path.exists(pasta):
        os.makedirs(pasta)

# CORES DHL
DHL_YELLOW = "#FFCC00"
DHL_RED = "#D40511"
BG_WHITE = "#FFFFFF"


# ==============================================================================
# --- SISTEMA DE LOG E AUDITORIA EM EXCEL ---
# ==============================================================================
class AuditLogger:
    @staticmethod
    def log(acao, detalhes="", arquivo_origem="", arquivo_destino="", status_final="", tempo_processamento=""):
        try:
            usuario = os.getlogin()
            maquina = socket.gethostname()
            data_hoje = datetime.datetime.now().strftime("%Y-%m-%d")
            caminho_log = os.path.join(DIR_LOGS, "Starlink_Audit_Log.xlsx")
            
            timestamp = datetime.datetime.now()
            
            # Criar ou carregar workbook
            if os.path.exists(caminho_log):
                wb = load_workbook(caminho_log)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Audit Log"
                
                # Criar cabeçalhos
                headers = ["Timestamp", "Usuário", "Máquina", "Ação", "Detalhes", "Arquivo Origem", 
                          "Arquivo Destino", "Status Final", "Tempo Processamento", "Partner", "Flow"]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
            
            # Encontrar próxima linha vazia
            next_row = ws.max_row + 1
            
            # Obter valores dos campos Partner e Flow das variáveis fixas
            partner_value = ""
            flow_value = ""
            try:
                if hasattr(AuditLogger, '_instance_ref') and AuditLogger._instance_ref:
                    partner_value = getattr(AuditLogger._instance_ref, 'PARTNER_VALUE', 'BR0194SPX')
                    flow_value = getattr(AuditLogger._instance_ref, 'FLOW_VALUE', 'STARLINK')
            except:
                partner_value = 'BR0194SPX'
                flow_value = 'STARLINK'
            
            # Adicionar dados
            dados = [
                timestamp,
                usuario,
                maquina,
                acao.upper(),
                detalhes,
                arquivo_origem,
                arquivo_destino,
                status_final,
                tempo_processamento,
                partner_value,
                flow_value
            ]
            
            for col, valor in enumerate(dados, 1):
                ws.cell(row=next_row, column=col, value=valor)
            
            # Ajustar largura das colunas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(caminho_log)
            
        except Exception as e:
            # Fallback para log simples se Excel falhar
            try:
                caminho_txt = os.path.join(DIR_LOGS, f"Audit_Error_{data_hoje}.txt")
                with open(caminho_txt, "a", encoding="utf-8") as f:
                    f.write(f"[{datetime.datetime.now()}] ERRO EXCEL: {str(e)} - {acao}: {detalhes}\n")
            except:
                pass

    @staticmethod
    def set_instance_ref(instance):
        """Define referência da instância principal para acessar campos Partner/Flow"""
        AuditLogger._instance_ref = instance


# ==============================================================================
# --- CLASSE PRINCIPAL ---
# ==============================================================================
class ConversorXML_Starlink:
    def __init__(self, root):
        self.root = root
        self.root.title("Starlink XML - Manager & Upload")
        
        # Configurar ícone personalizado
        try:
            logo_path = resource_path("logo.ico")
            if os.path.exists(logo_path):
                self.root.iconbitmap(logo_path)
        except:
            pass  # Se não encontrar o ícone, usa o padrão
            
        try:
            self.root.state('zoomed')
        except:
            self.root.attributes('-fullscreen', True)
        self.root.configure(bg=BG_WHITE)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        
        # Configura ESC para fechar janela principal
        self.root.bind('<Escape>', lambda e: self.on_close())

        # Configurar referência para AuditLogger acessar campos Partner/Flow
        AuditLogger.set_instance_ref(self)
        AuditLogger.log("SISTEMA", "Programa Iniciado", status_final="SUCESSO")

        self.dados_conversao = {}
        self.dados_upload = {}
        self.scan_ativo = False
        self.driver = None  # Instância do Selenium
        self.dhl_status = "🔄 INICIANDO..."  # Status do DHL Link
        self.monitoring_dhl = True  # Flag para monitoramento contínuo do DHL
        
        # Configurações fixas do DHL Link (não editáveis pelo usuário)
        self.PARTNER_VALUE = "BR0194SPX"
        self.FLOW_VALUE = "STARLINK"
        self.campos_ja_preenchidos = False  # Controla se já preencheu Partner/Flow uma vez

        self._criar_icones_dinamicamente()
        self._configurar_estilos()
        self._criar_layout()

        # Iniciar Navegador em Segundo Plano imediatamente para otimizar performance
        self.update_status("Iniciando navegador DHL Link em segundo plano...")
        threading.Thread(target=self.iniciar_navegador_background, daemon=True).start()
        
        # Iniciar monitoramento contínuo do status DHL Link
        self.root.after(5000, self.monitorar_dhl_status)

        # Scan inicial após um pequeno delay
        self.root.after(1000, self.iniciar_scan_conversao)

    def iniciar_navegador_background(self):
        """Abre o site DHL Link em modo minimizado e faz a autenticação inicial"""
        self.update_status("🌐 Configurando navegador Edge...")
        try:
            opcoes = Options()
            # Configurações simplificadas e mais compatíveis
            opcoes.add_argument("--start-maximized")
            opcoes.add_argument("--disable-blink-features=AutomationControlled")
            opcoes.add_experimental_option("excludeSwitches", ["enable-automation"])
            opcoes.add_experimental_option('useAutomationExtension', False)
            
            # Suprimir logs do navegador
            opcoes.add_argument("--log-level=3")  # Só erros críticos
            opcoes.add_argument("--disable-logging")
            opcoes.add_argument("--disable-gpu-logging")
            opcoes.add_argument("--silent")
            opcoes.add_experimental_option('excludeSwitches', ['enable-logging'])
            opcoes.add_experimental_option('useAutomationExtension', False)

            self.update_status("🚀 Iniciando Edge...")
            self.driver = webdriver.Edge(options=opcoes)
            
            # Remove a propriedade webdriver para não ser detectado
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

            self.update_status("🔗 Acessando DHL Link...")
            # Acessa diretamente a página inicial
            self.driver.get("https://link-cc.dhl.com/")
            
            # Minimiza após carregar a página
            time.sleep(2)
            self.driver.minimize_window()
            self.update_status("📱 Navegador minimizado - aguarde o login...")

            # Aguarda detecção de autenticação com timeout menor
            timeout = 180  # 3 minutos
            self.update_status(f"⏳ Aguardando autenticação (até {timeout//60} min)...")
            
            # Loop para verificar autenticação
            start_time = time.time()
            authenticated = False
            
            while time.time() - start_time < timeout:
                try:
                    # Verifica se saiu da página inicial ou se encontra elementos da página interna
                    current_url = self.driver.current_url
                    if "login" not in current_url.lower() and "auth" not in current_url.lower():
                        # Tenta encontrar elementos que indicam que está logado
                        elements = self.driver.find_elements(By.XPATH, "//a[contains(@href, 'upload')] | //button[contains(text(), 'Upload')] | //*[text()='UPLOAD']")
                        if elements or "/upload" in current_url or "dashboard" in current_url:
                            authenticated = True
                            break
                    time.sleep(5)  # Verifica a cada 5 segundos
                except:
                    time.sleep(5)
                    continue

            if authenticated:
                # Navega para a página de upload
                self.update_status("🔄 Direcionando para página de upload...")
                self.driver.get(URL_DHL_LINK)
                time.sleep(3)
                
                # Verifica se conseguiu acessar a página de upload
                if "/upload" in self.driver.current_url:
                    self.dhl_status = "✅ CONECTADO"
                    self.update_status("✅ Edge PRONTO! DHL Link conectado na página de upload")
                    
                    # Preenche campos Partner e Flow apenas na primeira vez
                    self.preencher_campos_dhl_primeira_vez()
                else:
                    self.dhl_status = "🔄 CARREGANDO..."
                    self.update_status("Autenticado, acessando página de upload...")
                    
                self.update_dhl_status_display()
                AuditLogger.log("SELENIUM", "Sessão Edge autenticada com sucesso")
            else:
                self.dhl_status = "AGUARDANDO"
                self.update_dhl_status_display()
                self.update_status("Timeout na autenticação - finalize o login manualmente")

        except Exception as e:
            self.dhl_status = "ERRO"
            self.update_dhl_status_display()
            self.update_status(f"ERRO: {str(e)[:60]}...")
            AuditLogger.log("ERRO_SELENIUM", f"Falha na inicialização do Edge: {str(e)}")
            self.driver = None

    def on_close(self):
        AuditLogger.log("SISTEMA", "Programa Fechado pelo Usuário")
        self.monitoring_dhl = False  # Para o monitoramento
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
        self.root.destroy()

    def _configurar_estilos(self):
        style = ttk.Style()
        style.theme_use('clam')
        # Estilos das tabelas (baseado no arquivo de referência)
        style.configure("Treeview", background=BG_WHITE, foreground="black", rowheight=45, font=('Segoe UI', 9),
                        fieldbackground=BG_WHITE)
        style.configure("Treeview.Heading", background="#2E8B57", foreground="white", font=('Segoe UI', 12, 'bold'),
                        relief="raised", borderwidth=1)
        style.map("Treeview.Heading", background=[('active', "#228B22")])
        # Remover efeitos de seleção e hover nas linhas
        style.map("Treeview", 
                  background=[('selected', BG_WHITE), ('focus', BG_WHITE), ('!focus', BG_WHITE)], 
                  foreground=[('selected', 'black'), ('focus', 'black'), ('!focus', 'black')])
        
        # Botões com design moderno simulando cantos arredondados
        style.configure("Red.TButton", 
                        font=('Segoe UI', 10, 'bold'), 
                        background=DHL_RED, 
                        foreground="white",
                        relief="flat", 
                        borderwidth=3, 
                        focuscolor='none', 
                        padding=(18, 14))
        style.map("Red.TButton", 
                  background=[('active', '#E74C3C'), ('pressed', '#C0392B')],
                  bordercolor=[('active', '#C0392B'), ('!active', DHL_RED)],
                  relief=[('pressed', 'groove'), ('!pressed', 'ridge')])
                  
        style.configure("White.TButton", 
                        font=('Segoe UI', 9, 'bold'), 
                        background='#F8F9FA', 
                        foreground='#2C3E50',
                        relief="flat", 
                        borderwidth=3, 
                        focuscolor='none', 
                        padding=(16, 12))
        style.map("White.TButton", 
                  background=[('active', '#E9ECEF'), ('pressed', '#DEE2E6')],
                  bordercolor=[('active', '#95A5A6'), ('!active', '#BDC3C7')],
                  relief=[('pressed', 'groove'), ('!pressed', 'ridge')])

        # Estilos das abas - aba ativa maior que inativas
        style.configure("TNotebook.Tab", 
                        font=('Segoe UI', 10, 'bold'), 
                        padding=[15, 8],
                        background="#D0D0D0",
                        foreground="#666")
        style.map("TNotebook.Tab", 
                  background=[("selected", DHL_YELLOW), ("active", "#E0E0E0")], 
                  foreground=[("selected", DHL_RED), ("active", "#333")],
                  padding=[("selected", [20, 12]), ("active", [17, 10])])  # Aba selecionada maior

    def _criar_icones_dinamicamente(self):
        size = 20
        self.img_unchecked = tk.PhotoImage(width=size, height=size)
        self.img_checked = tk.PhotoImage(width=size, height=size)
        row_data = "{" + " ".join(["#FFFFFF"] * size) + "} "
        full_data = row_data * size
        self.img_unchecked.put(full_data)
        self.img_checked.put(full_data)

        for i in range(size):
            for img in [self.img_unchecked, self.img_checked]:
                img.put("#444", (i, 0));
                img.put("#444", (i, size - 1))
                img.put("#444", (0, i));
                img.put("#444", (size - 1, i))

        check_color = "#008000"
        for i in range(4):
            self.img_checked.put(check_color, (5 + i, 11 + i));
            self.img_checked.put(check_color, (5 + i, 12 + i))
        for i in range(8):
            self.img_checked.put(check_color, (8 + i, 14 - i));
            self.img_checked.put(check_color, (8 + i, 15 - i))

    def _criar_layout(self):
        header = tk.Frame(self.root, bg=DHL_YELLOW, height=80)
        header.pack(fill=tk.X, side=tk.TOP)
        
        # Título à esquerda
        titulo_frame = tk.Frame(header, bg=DHL_YELLOW)
        titulo_frame.pack(side=tk.LEFT, padx=30, pady=20)
        tk.Label(titulo_frame, text="STARLINK XML - GERENCIADOR DE INTEGRAÇÃO", bg=DHL_YELLOW, fg=DHL_RED,
                 font=('Segoe UI', 18, 'bold', 'italic')).pack()

        # Pesquisa no centro com estilo do ImpressaoNF_Dev
        search_frame = tk.Frame(header, bg=DHL_YELLOW)
        search_frame.pack(side=tk.LEFT, padx=50, pady=20)
        
        # Canvas para o filtro estilizado
        self.canvas_search_header = tk.Canvas(search_frame, width=280, height=35, 
                                             bg=DHL_YELLOW, highlightthickness=0)
        self.canvas_search_header.pack()
        
        # Desenhar campo arredondado (estilo ImpressaoNF_Dev)
        self._round_rectangle(self.canvas_search_header, 3, 3, 277, 32, radius=6,
                              fill="#D0D0D0", outline="")
        self._round_rectangle(self.canvas_search_header, 1, 2, 275, 31, radius=6,
                              fill="white", outline="#ACACAC")
        
        # Variável de busca global
        self.var_busca_global = tk.StringVar()
        self.var_busca_global.trace("w", lambda *args: self.filtrar_aba_atual())
        
        # Ícone lupa
        try:
            search_icon_header = tk.Label(self.canvas_search_header, text="🔍", bg="white", 
                                         font=('Segoe UI', 12), fg="#666")
            self.canvas_search_header.create_window(18, 17, window=search_icon_header)
        except:
            self.canvas_search_header.create_text(18, 17, text="🔍", font=("Segoe UI", 12), 
                                                 fill="#666", anchor="center")
        
        # Campo de entrada
        self.entry_busca_global = tk.Entry(self.canvas_search_header, textvariable=self.var_busca_global, 
                                          width=25, font=('Segoe UI', 10), bg="white", bd=0, 
                                          highlightthickness=0)
        self.canvas_search_header.create_window(35, 17, window=self.entry_busca_global, anchor="w")
        
        # Botão limpar
        btn_clear = tk.Label(self.canvas_search_header, text="✕", bg="white", fg="#999", cursor="hand2",
                            font=("Arial", 10, "bold"))
        btn_clear.bind("<Button-1>", lambda e: self.var_busca_global.set(""))
        btn_clear.bind("<Enter>", lambda e: btn_clear.config(fg="#333"))
        btn_clear.bind("<Leave>", lambda e: btn_clear.config(fg="#999"))
        self.canvas_search_header.create_window(255, 17, window=btn_clear)

        # Status DHL Link no cabeçalho (direita)
        status_frame = tk.Frame(header, bg=DHL_YELLOW)
        status_frame.pack(side=tk.RIGHT, padx=30, pady=20)
        
        tk.Label(status_frame, text="DHL LINK:", bg=DHL_YELLOW, fg=DHL_RED,
                 font=('Segoe UI', 12, 'bold')).pack(side=tk.LEFT)
        self.lbl_dhl_status = tk.Label(status_frame, text=self.dhl_status, bg=DHL_YELLOW, fg=DHL_RED,
                                       font=('Segoe UI', 12, 'bold'))
        self.lbl_dhl_status.pack(side=tk.LEFT, padx=(10, 15))
        
        # Botão de reconexão (inicialmente oculto)
        self.btn_reconectar = tk.Button(status_frame, text="RECONECTAR", bg='#E74C3C', fg='white', 
                                       font=('Segoe UI', 9, 'bold'), cursor='hand2', relief='flat',
                                       command=self.reconectar_dhl_link, padx=10, pady=4,
                                       activebackground='#E74C3C', activeforeground='white',
                                       bd=2, highlightthickness=2, highlightcolor='#34495E')
        # Inicialmente não exibe o botão

        # NOTEBOOK (ABAS)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # ABA 1: CONVERSÃO
        self.tab_conv = tk.Frame(self.notebook, bg=BG_WHITE)
        self.notebook.add(self.tab_conv, text="  📝 CONVERSÃO XML  ")
        self._construir_aba_conversao()

        # ABA 2: UPLOAD
        self.tab_up = tk.Frame(self.notebook, bg=BG_WHITE)
        self.notebook.add(self.tab_up, text="  📤 UPLOAD DHL LINK  ")
        self._construir_aba_upload()

        # ABA 3: HISTÓRICO
        self.tab_hist = tk.Frame(self.notebook, bg=BG_WHITE)
        self.notebook.add(self.tab_hist, text="  📊 HISTÓRICO DE ENVIOS  ")
        self._construir_aba_historico()

        # Dados para histórico
        self.dados_historico = {}

        self.notebook.bind("<<NotebookTabChanged>>", self.ao_trocar_aba)

    def _construir_aba_conversao(self):
        cols = ("id", "status", "xml_origem", "xml_convertido")
        
        # Variável de busca local sincronizada com global
        self.var_busca_conv = tk.StringVar()
        self.var_busca_conv.trace("w", lambda *args: self.filtrar_conversao())
        
        # Container para Treeview e Scrollbar
        tree_container = tk.Frame(self.tab_conv, bg=BG_WHITE)
        tree_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Treeview com nova coluna de chave
        cols = ("id", "chave", "status", "xml_origem", "xml_convertido")
        self.tree_conv = ttk.Treeview(tree_container, columns=cols, show='tree headings', selectmode="browse")

        self.tree_conv.heading("#0", text="SEL")
        self.tree_conv.column("#0", width=60, stretch=False, anchor='center')
        self.tree_conv.heading("id", text="#")
        self.tree_conv.column("id", width=50, stretch=False, anchor='center')
        self.tree_conv.heading("chave", text="CHAVE NFe")
        self.tree_conv.column("chave", width=200, stretch=False, anchor='center')
        self.tree_conv.heading("status", text="STATUS")
        self.tree_conv.column("status", width=120, stretch=False, anchor='center')
        self.tree_conv.heading("xml_origem", text="ARQUIVO ORIGINAL")
        self.tree_conv.column("xml_origem", width=250, stretch=True, anchor='center')
        self.tree_conv.heading("xml_convertido", text="ARQUIVO CONVERTIDO")
        self.tree_conv.column("xml_convertido", width=250, stretch=True, anchor='center')

        self.tree_conv.tag_configure("checked", image=self.img_checked)
        self.tree_conv.tag_configure("unchecked", image=self.img_unchecked)
        self.tree_conv.tag_configure("erro", foreground=DHL_RED, font=('Segoe UI', 9, 'bold'))
        self.tree_conv.tag_configure("concluido", foreground=DHL_RED, font=('Segoe UI', 9, 'bold'))  # Texto vermelho para convertidos

        # Scrollbar
        scroll = AutoScrollbar(tree_container, orient="vertical", command=self.tree_conv.yview)
        self.tree_conv.configure(yscrollcommand=scroll.set)
        
        # Pack correto: scrollbar à direita, treeview preenchendo o resto
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree_conv.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.tree_conv.bind('<Button-1>', lambda e: self.on_click_tree(e, self.tree_conv, self.dados_conversao))
        self.tree_conv.bind('<Double-1>', lambda e: self.abrir_visualizador(e, self.tree_conv, self.dados_conversao))
        # Remover seleção visual
        self.tree_conv.bind('<FocusIn>', lambda e: self.tree_conv.selection_remove(self.tree_conv.selection()))

        # Informativo sobre clique duplo
        info_frame = tk.Frame(self.tab_conv, bg=BG_WHITE, height=25)
        info_frame.pack(fill=tk.X, padx=10, pady=(0, 5))
        tk.Label(info_frame, text="💡 Duplo clique para visualizar XML | Busca por chave | Logs automáticos", 
                bg=BG_WHITE, fg="#666", font=('Segoe UI', 9, 'italic')).pack(side=tk.LEFT)

        # Linha de status da aba (entre tabela e botões)
        self.lbl_status_conv = tk.Label(self.tab_conv, text=" 🔄 Sistema inicializado - Aguardando arquivos XML", 
                                       bg='#34495E', fg='#ECF0F1', anchor='w',
                                       font=('Segoe UI', 10, 'bold'), height=1, relief='flat')
        self.lbl_status_conv.pack(fill=tk.X, padx=10, pady=(5, 10))

        bot = tk.Frame(self.tab_conv, bg=DHL_YELLOW, height=70)
        bot.pack(fill=tk.X, side=tk.BOTTOM)

        ttk.Button(bot, text="ATUALIZAR LISTA", command=self.iniciar_scan_conversao, style="White.TButton").pack(
            side=tk.LEFT, padx=20, pady=15)
        self.btn_toggle_conv = ttk.Button(bot, text="MARCAR TODOS", command=lambda: self.toggle_all(self.tree_conv),
                                          style="White.TButton", width=15)
        self.btn_toggle_conv.pack(side=tk.LEFT, padx=5, pady=15)
        self.lbl_resumo_conv = tk.Label(bot, text="SELECIONADOS: 0", bg=DHL_YELLOW, fg=DHL_RED,
                                        font=('Segoe UI', 12, 'bold'))
        self.lbl_resumo_conv.pack(side=tk.LEFT, padx=30)

        # Botão principal com cantos arredondados
        self.btn_acao_conv = self.criar_botao_arredondado(bot, "CONVERTER AGORA", self.executar_conversao,
                                                         DHL_RED, "white", width=160, height=45)
        self.btn_acao_conv.pack(side=tk.RIGHT, padx=20, pady=12)

    def _construir_aba_upload(self):
        cols = ("id", "status", "xml_convertido")
        
        # Variável de busca local sincronizada com global
        self.var_busca_upload = tk.StringVar()
        self.var_busca_upload.trace("w", lambda *args: self.filtrar_upload())
        
        # Container para Treeview e Scrollbar  
        tree_frame = tk.Frame(self.tab_up, bg=BG_WHITE)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Configurar grid no container
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        self.tree_up = ttk.Treeview(tree_frame, columns=cols, show='tree headings', selectmode="browse")

        self.tree_up.heading("#0", text="SEL")
        self.tree_up.column("#0", width=60, stretch=False, anchor='center')
        self.tree_up.heading("id", text="#")
        self.tree_up.column("id", width=50, stretch=False, anchor='center')
        self.tree_up.heading("status", text="STATUS DO UPLOAD")
        self.tree_up.column("status", width=250, stretch=False, anchor='center')
        self.tree_up.heading("xml_convertido", text="ARQUIVO PRONTO (XML BASE64)")
        self.tree_up.column("xml_convertido", width=400, stretch=True, anchor='center')

        self.tree_up.tag_configure("checked", image=self.img_checked)
        self.tree_up.tag_configure("unchecked", image=self.img_unchecked)
        self.tree_up.tag_configure("erro", foreground=DHL_RED, font=('', 9, 'bold'))
        self.tree_up.tag_configure("sucesso", foreground="#008000", font=('', 9, 'bold'))

        # Scrollbar
        scroll = AutoScrollbar(tree_frame, orient="vertical", command=self.tree_up.yview)
        self.tree_up.configure(yscrollcommand=scroll.set)
        
        # Pack correto: scrollbar à direita, treeview preenchendo o resto
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree_up.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.tree_up.bind('<Button-1>', lambda e: self.on_click_tree(e, self.tree_up, self.dados_upload))
        # Remover seleção visual
        self.tree_up.bind('<FocusIn>', lambda e: self.tree_up.selection_remove(self.tree_up.selection()))
        self.tree_up.bind('<Double-1>', lambda e: self.abrir_visualizador(e, self.tree_up, self.dados_upload))

        # Informativo sobre clique duplo
        info_frame = tk.Frame(self.tab_up, bg=BG_WHITE, height=25)
        info_frame.pack(fill=tk.X, padx=10, pady=(0, 5))
        tk.Label(info_frame, text=f"Configuração: {self.PARTNER_VALUE} | {self.FLOW_VALUE} | Envios em lotes de 100 arquivos", 
                bg=BG_WHITE, fg="#666", font=('Segoe UI', 9, 'italic')).pack(side=tk.LEFT)

        # Linha de status da aba (entre tabela e botões)
        self.lbl_status_upload = tk.Label(self.tab_up, text=" 🔄 Inicializando conexão DHL Link - Aguarde...", 
                                         bg='#34495E', fg='#ECF0F1', anchor='w',
                                         font=('Segoe UI', 10, 'bold'), height=1, relief='flat')
        self.lbl_status_upload.pack(fill=tk.X, padx=10, pady=(5, 10))

        bot = tk.Frame(self.tab_up, bg=DHL_YELLOW, height=70)
        bot.pack(fill=tk.X, side=tk.BOTTOM)

        ttk.Button(bot, text="ATUALIZAR LISTA", command=self.iniciar_scan_upload, style="White.TButton").pack(
            side=tk.LEFT, padx=20, pady=15)
        self.btn_toggle_up = ttk.Button(bot, text="MARCAR TODOS", command=lambda: self.toggle_all(self.tree_up),
                                        style="White.TButton", width=20)
        self.btn_toggle_up.pack(side=tk.LEFT, padx=5, pady=15)
        self.lbl_resumo_up = tk.Label(bot, text="SELECIONADOS: 0", bg=DHL_YELLOW, fg=DHL_RED,
                                      font=('Segoe UI', 12, 'bold'))
        self.lbl_resumo_up.pack(side=tk.LEFT, padx=30)

        # Botão principal com cantos arredondados
        self.btn_acao_up = self.criar_botao_arredondado(bot, "ENVIAR PARA DHL LINK", self.executar_upload,
                                                       DHL_RED, "white", width=180, height=45)
        self.btn_acao_up.pack(side=tk.RIGHT, padx=20, pady=12)

    def _construir_aba_historico(self):
        """Constrói a aba de histórico de arquivos enviados"""
        cols = ("id", "data_envio", "arquivo", "data_exclusao", "status_envio")
        
        # Variável de busca local sincronizada com global
        self.var_busca_hist = tk.StringVar()
        self.var_busca_hist.trace("w", lambda *args: self.filtrar_historico())
        
        # Container para Treeview e Scrollbar
        tree_container = tk.Frame(self.tab_hist, bg=BG_WHITE)
        tree_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Configurar grid no container
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)
        
        # Treeview
        self.tree_hist = ttk.Treeview(tree_container, columns=cols, show='tree headings', selectmode="browse")

        self.tree_hist.heading("#0", text="SEL")
        self.tree_hist.column("#0", width=60, stretch=False, anchor='center')
        self.tree_hist.heading("id", text="#")
        self.tree_hist.column("id", width=50, stretch=False, anchor='center')
        self.tree_hist.heading("data_envio", text="DATA ENVIO")
        self.tree_hist.column("data_envio", width=120, stretch=False, anchor='center')
        self.tree_hist.heading("arquivo", text="ARQUIVO ENVIADO")
        self.tree_hist.column("arquivo", width=350, stretch=True, anchor='center')
        self.tree_hist.heading("data_exclusao", text="EXCLUSÃO EM")
        self.tree_hist.column("data_exclusao", width=120, stretch=False, anchor='center')
        self.tree_hist.heading("status_envio", text="STATUS")
        self.tree_hist.column("status_envio", width=100, stretch=False, anchor='center')

        self.tree_hist.tag_configure("checked", image=self.img_checked)
        self.tree_hist.tag_configure("unchecked", image=self.img_unchecked)
        self.tree_hist.tag_configure("sucesso", foreground="#008000", font=('Segoe UI', 9, 'bold'))
        self.tree_hist.tag_configure("erro", foreground=DHL_RED, font=('Segoe UI', 9, 'bold'))

        # Scrollbar
        scroll = AutoScrollbar(tree_container, orient="vertical", command=self.tree_hist.yview)
        self.tree_hist.configure(yscrollcommand=scroll.set)
        
        # Pack correto: scrollbar à direita, treeview preenchendo o resto
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree_hist.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.tree_hist.bind('<Button-1>', lambda e: self.on_click_tree(e, self.tree_hist, self.dados_historico))
        self.tree_hist.bind('<FocusIn>', lambda e: self.tree_hist.selection_remove(self.tree_hist.selection()))
        self.tree_hist.bind('<Double-1>', lambda e: self.abrir_visualizador_historico(e))

        # Informativo sobre histórico
        info_frame = tk.Frame(self.tab_hist, bg=BG_WHITE, height=25)
        info_frame.pack(fill=tk.X, padx=10, pady=(0, 5))
        tk.Label(info_frame, text="💡 Arquivos são excluídos automaticamente 2 dias após envio | Duplo clique para visualizar XML | Reenvio disponível", 
                bg=BG_WHITE, fg="#666", font=('Segoe UI', 9, 'italic')).pack(side=tk.LEFT)

        # Linha de status da aba
        self.lbl_status_hist = tk.Label(self.tab_hist, text=" 📊 Carregando histórico de envios...", 
                                       bg='#34495E', fg='#ECF0F1', anchor='w',
                                       font=('Segoe UI', 10, 'bold'), height=1, relief='flat')
        self.lbl_status_hist.pack(fill=tk.X, padx=10, pady=(5, 10))

        # Botões de ação
        bot = tk.Frame(self.tab_hist, bg=DHL_YELLOW, height=70)
        bot.pack(fill=tk.X, side=tk.BOTTOM)

        ttk.Button(bot, text="ATUALIZAR HISTÓRICO", command=self.carregar_historico, style="White.TButton").pack(
            side=tk.LEFT, padx=20, pady=15)
        self.btn_toggle_hist = ttk.Button(bot, text="MARCAR TODOS", command=lambda: self.toggle_all(self.tree_hist),
                                         style="White.TButton", width=15)
        self.btn_toggle_hist.pack(side=tk.LEFT, padx=5, pady=15)
        self.lbl_resumo_hist = tk.Label(bot, text="SELECIONADOS: 0", bg=DHL_YELLOW, fg=DHL_RED,
                                       font=('Segoe UI', 12, 'bold'))
        self.lbl_resumo_hist.pack(side=tk.LEFT, padx=30)

        # Botão principal reenviar
        self.btn_acao_hist = self.criar_botao_arredondado(bot, "REENVIAR SELECIONADOS", self.executar_reenvio,
                                                         "#E67E22", "white", width=200, height=45)
        self.btn_acao_hist.pack(side=tk.RIGHT, padx=20, pady=12)

    def update_status(self, msg):
        """Atualiza o status na aba ativa"""
        try:
            aba_atual = self.notebook.index("current")
            if aba_atual == 0 and hasattr(self, 'lbl_status_conv'):
                self.lbl_status_conv.config(text=f" {msg}")
            elif aba_atual == 1 and hasattr(self, 'lbl_status_upload'):
                self.lbl_status_upload.config(text=f" {msg}")
            self.root.update_idletasks()
        except:
            pass
        
    def configurar_icone_janela(self, janela):
        """Configura o ícone personalizado para uma janela"""
        try:
            logo_path = resource_path("logo.ico")
            if os.path.exists(logo_path):
                janela.iconbitmap(logo_path)
        except:
            pass
    
    def configurar_tecla_esc(self, janela):
        """Configura a tecla ESC para fechar a janela"""
        def fechar_com_esc(event):
            janela.destroy()
        
        janela.bind('<Escape>', fechar_com_esc)
        janela.focus_set()  # Garante que a janela tenha foco para receber eventos de teclado
    
    def criar_botao_arredondado(self, parent, text, command, bg_color, text_color, width=120, height=40):
        """Cria um botão com cantos arredondados usando Canvas"""
        canvas = tk.Canvas(parent, width=width, height=height, highlightthickness=0, 
                          relief='flat', borderwidth=0, bg=parent.cget('bg'))
        
        # Desenhar retângulo arredondado
        radius = 8
        x1, y1, x2, y2 = 2, 2, width-2, height-2
        
        # Fundo do botão
        canvas.create_oval(x1, y1, x1+radius*2, y1+radius*2, fill=bg_color, outline=bg_color)
        canvas.create_oval(x2-radius*2, y1, x2, y1+radius*2, fill=bg_color, outline=bg_color)
        canvas.create_oval(x1, y2-radius*2, x1+radius*2, y2, fill=bg_color, outline=bg_color)
        canvas.create_oval(x2-radius*2, y2-radius*2, x2, y2, fill=bg_color, outline=bg_color)
        
        canvas.create_rectangle(x1+radius, y1, x2-radius, y2, fill=bg_color, outline=bg_color)
        canvas.create_rectangle(x1, y1+radius, x2, y2-radius, fill=bg_color, outline=bg_color)
        
        # Texto do botão
        canvas.create_text(width//2, height//2, text=text, fill=text_color, 
                          font=('Segoe UI', 10, 'bold'), anchor='center')
        
        # Eventos de clique
        def on_click(event):
            if command:
                command()
        
        def on_enter(event):
            # Efeito hover - escurece a cor
            darker_color = self.escurecer_cor(bg_color, 0.1)
            canvas.delete("all")
            # Redesenhar com cor mais escura
            canvas.create_oval(x1, y1, x1+radius*2, y1+radius*2, fill=darker_color, outline=darker_color)
            canvas.create_oval(x2-radius*2, y1, x2, y1+radius*2, fill=darker_color, outline=darker_color)
            canvas.create_oval(x1, y2-radius*2, x1+radius*2, y2, fill=darker_color, outline=darker_color)
            canvas.create_oval(x2-radius*2, y2-radius*2, x2, y2, fill=darker_color, outline=darker_color)
            canvas.create_rectangle(x1+radius, y1, x2-radius, y2, fill=darker_color, outline=darker_color)
            canvas.create_rectangle(x1, y1+radius, x2, y2-radius, fill=darker_color, outline=darker_color)
            canvas.create_text(width//2, height//2, text=text, fill=text_color, 
                              font=('Segoe UI', 10, 'bold'), anchor='center')
        
        def on_leave(event):
            # Volta cor original
            canvas.delete("all")
            canvas.create_oval(x1, y1, x1+radius*2, y1+radius*2, fill=bg_color, outline=bg_color)
            canvas.create_oval(x2-radius*2, y1, x2, y1+radius*2, fill=bg_color, outline=bg_color)
            canvas.create_oval(x1, y2-radius*2, x1+radius*2, y2, fill=bg_color, outline=bg_color)
            canvas.create_oval(x2-radius*2, y2-radius*2, x2, y2, fill=bg_color, outline=bg_color)
            canvas.create_rectangle(x1+radius, y1, x2-radius, y2, fill=bg_color, outline=bg_color)
            canvas.create_rectangle(x1, y1+radius, x2, y2-radius, fill=bg_color, outline=bg_color)
            canvas.create_text(width//2, height//2, text=text, fill=text_color, 
                              font=('Segoe UI', 10, 'bold'), anchor='center')
        
        canvas.bind("<Button-1>", on_click)
        canvas.bind("<Enter>", on_enter)
        canvas.bind("<Leave>", on_leave)
        canvas.configure(cursor="hand2")
        
        return canvas
    
    def escurecer_cor(self, cor, fator):
        """Escurece uma cor em formato hex"""
        if cor.startswith('#'):
            cor = cor[1:]
        r, g, b = tuple(int(cor[i:i+2], 16) for i in (0, 2, 4))
        r = int(r * (1 - fator))
        g = int(g * (1 - fator))
        b = int(b * (1 - fator))
        return f"#{r:02x}{g:02x}{b:02x}"

    def criar_filtro_pesquisa(self, parent, var_busca, callback_filtro):
        """Cria um filtro de pesquisa visual baseado no arquivo de referência"""
        frame_busca = tk.Frame(parent, bg=BG_WHITE, height=50)
        frame_busca.pack(fill=tk.X, padx=10, pady=(10, 5))
        frame_busca.pack_propagate(False)
        
        tk.Label(frame_busca, text="🔍 Buscar por Chave:", bg=BG_WHITE, fg="#666",
                 font=('Segoe UI', 9, 'bold')).pack(anchor='w', pady=(8, 2))
        
        # Canvas para o campo de busca
        canvas_search = tk.Canvas(frame_busca, width=400, height=35, bg=BG_WHITE, highlightthickness=0)
        canvas_search.pack(anchor="w")
        
        # Desenhar campo arredondado
        self._round_rectangle(canvas_search, 2, 2, 398, 33, radius=6, fill="#F0F0F0", outline="#CCC")
        
        # Ícone de pesquisa
        try:
            search_icon = tk.Label(canvas_search, text="🔍", bg="#F0F0F0", font=('Segoe UI', 12))
            canvas_search.create_window(20, 17, window=search_icon)
        except:
            canvas_search.create_text(20, 17, text="🔍", font=("Segoe UI", 12), fill="#666", anchor="center")
        
        # Campo de entrada
        entry_busca = tk.Entry(canvas_search, textvariable=var_busca, width=35,
                              font=('Segoe UI', 10), bg="#F0F0F0", bd=0, highlightthickness=0)
        canvas_search.create_window(40, 17, window=entry_busca, anchor="w")
        
        # Botão limpar
        lbl_clear = tk.Label(canvas_search, text="✕", bg="#F0F0F0", fg="#999",
                            font=("Arial", 10, "bold"), cursor="hand2")
        lbl_clear.bind("<Button-1>", lambda e: [var_busca.set(""), callback_filtro()])
        canvas_search.create_window(370, 17, window=lbl_clear)
        
        return canvas_search

    def _round_rectangle(self, canvas, x1, y1, x2, y2, radius=25, **kwargs):
        """Desenha retângulo com cantos arredondados no Canvas (padrão ImpressaoNF_Dev)"""
        points = [x1 + radius, y1,
                  x1 + radius, y1,
                  x2 - radius, y1,
                  x2 - radius, y1,
                  x2, y1,
                  x2, y1 + radius,
                  x2, y1 + radius,
                  x2, y2 - radius,
                  x2, y2 - radius,
                  x2, y2,
                  x2 - radius, y2,
                  x2 - radius, y2,
                  x1 + radius, y2,
                  x1 + radius, y2,
                  x1, y2,
                  x1, y2 - radius,
                  x1, y2 - radius,
                  x1, y1 + radius,
                  x1, y1 + radius,
                  x1, y1]
        return canvas.create_polygon(points, smooth=True, **kwargs)

    def filtrar_aba_atual(self):
        """Filtra a aba atualmente ativa com base na pesquisa global"""
        try:
            aba_atual = self.notebook.index("current")
            termo = self.var_busca_global.get()
            
            if aba_atual == 0:  # Conversão
                if hasattr(self, 'var_busca_conv'):
                    self.var_busca_conv.set(termo)
            elif aba_atual == 1:  # Upload
                if hasattr(self, 'var_busca_upload'):
                    self.var_busca_upload.set(termo)
            elif aba_atual == 2:  # Histórico
                if hasattr(self, 'var_busca_hist'):
                    self.var_busca_hist.set(termo)
        except:
            pass

    def update_dhl_status_display(self):
        """Atualiza apenas o display do status DHL no cabeçalho"""
        if hasattr(self, 'lbl_dhl_status'):
            self.lbl_dhl_status.config(text=self.dhl_status)
            
            # Mostra/esconde botão de reconexão baseado no status
            if hasattr(self, 'btn_reconectar'):
                if "❌" in self.dhl_status or "⚠️" in self.dhl_status or "⏳" in self.dhl_status:
                    self.btn_reconectar.pack(side=tk.LEFT, padx=(0, 5))
                else:
                    self.btn_reconectar.pack_forget()
            
            self.root.update_idletasks()

    def monitorar_dhl_status(self):
        """Monitora constantemente o status do navegador DHL Link"""
        if not self.monitoring_dhl:
            return
            
        try:
            if self.driver:
                # Verifica se o navegador ainda está aberto e funcionando
                try:
                    current_url = self.driver.current_url
                    if "link-cc.dhl.com" in current_url:
                        if "/upload" in current_url:
                            if self.dhl_status != "✅ CONECTADO":
                                self.dhl_status = "✅ CONECTADO"
                                self.update_dhl_status_display()
                                if hasattr(self, 'lbl_status_upload'):
                                    self.lbl_status_upload.config(text=" ✅ DHL Link pronto para envio dos arquivos XML.")
                        else:
                            if self.dhl_status != "🔄 CARREGANDO...":
                                self.dhl_status = "🔄 CARREGANDO..."
                                self.update_dhl_status_display()
                                if hasattr(self, 'lbl_status_upload'):
                                    self.lbl_status_upload.config(text="DHL Link logado mas não na página de upload")
                    else:
                        if self.dhl_status not in ["⏳ AGUARDANDO", "🔄 INICIANDO..."]:
                            self.dhl_status = "⏳ AGUARDANDO"
                            self.update_dhl_status_display()
                            if hasattr(self, 'lbl_status_upload'):
                                self.lbl_status_upload.config(text=" ⏳ Aguardando login no DHL Link")
                except:
                    # Driver foi fechado ou não responde
                    self.dhl_status = "❌ DESCONECTADO"
                    self.update_dhl_status_display()
                    if hasattr(self, 'lbl_status_upload'):
                        self.lbl_status_upload.config(text=" ❌ Navegador DHL Link foi fechado")
                    self.driver = None
            else:
                # Driver é None, navegador não está disponível
                if self.dhl_status not in ["🔄 INICIANDO...", "❌ ERRO"]:
                    self.dhl_status = "❌ DESCONECTADO"
                    self.update_dhl_status_display()
                    if hasattr(self, 'lbl_status_upload'):
                        self.lbl_status_upload.config(text=" ❌ Navegador DHL Link não está disponível")
                        
        except Exception:
            pass  # Silenciosamente ignora erros de monitoramento
            
        # Reagenda o monitoramento para daqui a 10 segundos
        if self.monitoring_dhl:
            self.root.after(10000, self.monitorar_dhl_status)

    def reconectar_dhl_link(self):
        """Tenta reconectar ao DHL Link"""
        if hasattr(self, 'btn_reconectar'):
            self.btn_reconectar.config(state='disabled', text='🔄 CONECTANDO...')
        
        # Para o monitoramento temporariamente
        self.monitoring_dhl = False
        
        # Fecha o driver atual se existir
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
            self.driver = None
        
        # Atualiza status
        self.dhl_status = "🔄 RECONECTANDO..."
        self.update_dhl_status_display()
        
        # Inicia nova conexão em thread separada
        def reconectar():
            self.monitoring_dhl = True
            self.iniciar_navegador_background()
            
            # Reabilita botão após tentativa
            if hasattr(self, 'btn_reconectar'):
                self.btn_reconectar.config(state='normal', text='🔄 RECONECTAR')
        
        threading.Thread(target=reconectar, daemon=True).start()

    def filtrar_conversao(self):
        """Filtra a tabela de conversão por chave"""
        termo = self.var_busca_conv.get().strip().upper()
        
        # Limpa a tabela
        for item in self.tree_conv.get_children():
            self.tree_conv.delete(item)
        
        # Se não há termo, mostra todos
        if not termo:
            self.repopular_tree_conversao()
            return
        
        # Filtra por chave que contenha o termo
        for item_id, dados in self.dados_conversao.items():
            chave = dados.get('chave', '').upper()
            nome_arquivo = dados.get('original', '').upper()
            
            if termo in chave or termo in nome_arquivo:
                status = dados.get('status', 'Pendente')
                tags = ["unchecked"]
                
                if status == "Convertido":
                    tags = ["unchecked", "concluido"]
                elif status == "ERRO":
                    tags.append("erro")
                
                self.tree_conv.insert("", "end", iid=item_id, tags=tuple(tags),
                                     values=(item_id, chave[:20] + "..." if len(chave) > 20 else chave, 
                                           status, dados.get('original', ''), 
                                           dados.get('xml_convertido', '---')))

    def repopular_tree_conversao(self):
        """Repopula a árvore de conversão sem filtro"""
        for item_id, dados in self.dados_conversao.items():
            status = dados.get('status', 'Pendente')
            tags = ["unchecked"]
            
            if status == "Convertido":
                tags = ["unchecked", "concluido"]
            elif status == "ERRO":
                tags.append("erro")
            
            chave = dados.get('chave', '')
            self.tree_conv.insert("", "end", iid=item_id, tags=tuple(tags),
                                 values=(item_id, chave[:20] + "..." if len(chave) > 20 else chave, 
                                       status, dados.get('original', ''), 
                                       dados.get('xml_convertido', '---')))

    def preencher_campos_dhl_primeira_vez(self):
        """Preenche campos Partner e Flow apenas na primeira vez (histórico do navegador salva depois)"""
        if self.campos_ja_preenchidos:
            self.update_status(f"ℹ️ Campos Partner/Flow já configurados - usando histórico do navegador")
            return
            
        try:
            # Aguarda um pouco para a página carregar completamente
            time.sleep(3)
            
            # Preenche campo Partner apenas na primeira vez
            try:
                partner_field = self.driver.find_element(By.ID, "uploadFlowSelection_pl_flow_selector_web_pl_pl-search-text-input")
                partner_field.clear()
                partner_field.send_keys(self.PARTNER_VALUE)
                time.sleep(1)
            except:
                pass
            
            # Preenche campo Flow apenas na primeira vez
            try:
                flow_field = self.driver.find_element(By.ID, "uploadFlowSelection_pl_flow_selector_web_flow_flow-search-text-input")
                flow_field.clear()
                flow_field.send_keys(self.FLOW_VALUE)
                time.sleep(1)
            except:
                pass
            
            # Marca como já preenchido para não repetir
            self.campos_ja_preenchidos = True
            self.update_status(f"🔧 Configuração inicial aplicada: {self.PARTNER_VALUE} | {self.FLOW_VALUE}")
            
        except Exception as e:
            pass  # Silenciosamente ignora erros de preenchimento

    def preencher_campos_dhl(self):
        """Método legado - mantido para compatibilidade"""
        self.preencher_campos_dhl_primeira_vez()

    def ao_trocar_aba(self, event):
        idx = self.notebook.index("current")
        if idx == 0:
            self.iniciar_scan_conversao()
        elif idx == 1:
            self.iniciar_scan_upload()
        elif idx == 2:
            self.carregar_historico()

    def calcular_resumo(self, tree, label, btn_toggle):
        total = 0
        todos = True
        items = tree.get_children()
        if not items: todos = False
        for item in items:
            if "checked" in tree.item(item, "tags"):
                total += 1
            else:
                todos = False
        label.config(text=f"SELECIONADOS: {total}")
        btn_toggle.config(text="DESMARCAR TODOS" if (todos and items) else "MARCAR TODOS")

    def on_click_tree(self, event, tree, dados_dict):
        if tree.identify("region", event.x, event.y) == "heading": return
        row_id = tree.identify_row(event.y)
        if row_id and tree.identify_column(event.x) == "#0":
            tags = list(tree.item(row_id, "tags"))
            if "checked" in tags:
                tags.remove("checked"); tags.append("unchecked")
            else:
                tags.remove("unchecked"); tags.append("checked")
            tree.item(row_id, tags=tuple(tags))

            if tree == self.tree_conv:
                self.calcular_resumo(self.tree_conv, self.lbl_resumo_conv, self.btn_toggle_conv)
            else:
                self.calcular_resumo(self.tree_up, self.lbl_resumo_up, self.btn_toggle_up)
            
            # Remove completamente qualquer seleção visual
            tree.selection_remove(row_id)
            tree.focus_set()
            tree.focus('')

    def toggle_all(self, tree):
        items = tree.get_children()
        if not items: return

        if tree == self.tree_conv:
            btn = self.btn_toggle_conv; lbl = self.lbl_resumo_conv
        else:
            btn = self.btn_toggle_up; lbl = self.lbl_resumo_up

        novo_status = "checked" if btn.cget("text") == "MARCAR TODOS" else "unchecked"
        for item in items:
            tags = [t for t in tree.item(item, "tags") if t not in ["checked", "unchecked"]]
            tags.append(novo_status)
            tree.item(item, tags=tuple(tags))
        self.calcular_resumo(tree, lbl, btn)

    # ==========================================================================
    # --- LÓGICA: ABA 1 (CONVERSÃO XML) ---
    # ==========================================================================
    def iniciar_scan_conversao(self):
        if self.scan_ativo: return
        self.update_status("🔍 Buscando XMLs na pasta de origem...")
        for item in self.tree_conv.get_children(): self.tree_conv.delete(item)
        self.dados_conversao.clear()

        arquivos = [f for f in os.listdir(DIR_ENTRADA) if f.lower().endswith('.xml')]
        for i, f in enumerate(arquivos, 1):
            item_id = str(i)
            # Extrair chave do arquivo
            chave, _, _, _ = self.extrair_dados_nfe(os.path.join(DIR_ENTRADA, f))
            if not chave or chave == "CHAVE_NAO_ENCONTRADA":
                chave = f.split('.')[0]  # Usa nome do arquivo como fallback
                
            self.dados_conversao[item_id] = {
                'original': f, 
                'path_origem': os.path.join(DIR_ENTRADA, f),
                'status': "Pendente", 
                'path_convertido': "",
                'chave': chave
            }
            
            chave_display = chave[:20] + "..." if len(chave) > 20 else chave
            self.tree_conv.insert("", "end", iid=item_id, tags=("checked",),
                                  values=(i, chave_display, "Pendente", f, "---"))

        self.calcular_resumo(self.tree_conv, self.lbl_resumo_conv, self.btn_toggle_conv)
        total_arquivos = len(arquivos)
        if total_arquivos > 0:
            msg_plural = pluralizar(total_arquivos, "XML", "XMLs")
            status_msg = "Encontrado" if total_arquivos == 1 else "Encontrados"
            prontos_msg = "pronto" if total_arquivos == 1 else "prontos"
            self.update_status(f"✅ {status_msg} {msg_plural} {prontos_msg} para conversão")
            if hasattr(self, 'lbl_status_conv'):
                self.lbl_status_conv.config(text=f" ✅ {msg_plural} {prontos_msg} para conversão")
        else:
            self.update_status("ℹ️ Nenhum XML encontrado na pasta de origem")
            if hasattr(self, 'lbl_status_conv'):
                self.lbl_status_conv.config(text=" ℹ️ Nenhum XML encontrado na pasta de origem")

    def carregar_historico(self):
        """Carrega histórico de arquivos enviados do log Excel"""
        self.update_status("📊 Carregando histórico de envios...")
        for item in self.tree_hist.get_children(): 
            self.tree_hist.delete(item)
        self.dados_historico.clear()

        try:
            caminho_log = os.path.join(DIR_LOGS, "Starlink_Audit_Log.xlsx")
            if not os.path.exists(caminho_log):
                self.lbl_status_hist.config(text=" ℹ️ Nenhum histórico encontrado - arquivo de log não existe")
                return

            wb = load_workbook(caminho_log)
            ws = wb.active
            
            historico_count = 0
            
            # Ler dados do Excel (pular cabeçalho)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
                if not row or len(row) < 5:
                    continue
                    
                timestamp, usuario, maquina, acao, detalhes, arquivo_origem, arquivo_destino, status_final, tempo, partner, flow = row[:11]
                
                # Filtrar apenas uploads bem-sucedidos
                if acao == "UPLOAD" and status_final == "ENVIADO" and arquivo_origem:
                    historico_count += 1
                    item_id = str(historico_count)
                    
                    data_envio = timestamp.strftime("%d/%m/%Y %H:%M") if hasattr(timestamp, 'strftime') else str(timestamp)[:16]
                    
                    # Calcular data de exclusão (2 dias após envio)
                    if hasattr(timestamp, 'strftime'):
                        data_exclusao_obj = timestamp + datetime.timedelta(days=2)
                        data_exclusao = data_exclusao_obj.strftime("%d/%m/%Y %H:%M")
                        # Verificar se já passou da data de exclusão
                        agora = datetime.datetime.now()
                        ja_excluido = agora > data_exclusao_obj
                    else:
                        data_exclusao = "N/A"
                        ja_excluido = False
                    
                    self.dados_historico[item_id] = {
                        'arquivo_origem': arquivo_origem,
                        'data_envio': data_envio,
                        'data_exclusao': data_exclusao,
                        'status_envio': status_final,
                        'detalhes': detalhes,
                        'ja_excluido': ja_excluido
                    }
                    
                    # Definir tags baseadas no status de exclusão
                    tags = ["unchecked"]
                    if ja_excluido:
                        tags.append("erro")  # Vermelho para arquivos já excluídos
                        status_display = "EXCLUÍDO"
                    else:
                        tags.append("sucesso")
                        status_display = status_final
                    
                    self.tree_hist.insert("", "end", iid=item_id, tags=tuple(tags),
                                         values=(historico_count, data_envio, arquivo_origem, 
                                               data_exclusao, status_display))

            self.calcular_resumo(self.tree_hist, self.lbl_resumo_hist, self.btn_toggle_hist)
            
            if historico_count > 0:
                msg_plural = pluralizar(historico_count, "envio", "envios")
                self.lbl_status_hist.config(text=f" 📊 {msg_plural} encontrados no histórico")
            else:
                self.lbl_status_hist.config(text=" ℹ️ Nenhum envio encontrado no histórico")
                
        except Exception as e:
            self.lbl_status_hist.config(text=f" ❌ Erro ao carregar histórico: {str(e)[:50]}...")

    def filtrar_historico(self):
        """Filtra o histórico por nome de arquivo"""
        termo = self.var_busca_hist.get().strip().upper()
        
        # Limpa a tabela
        for item in self.tree_hist.get_children():
            self.tree_hist.delete(item)
        
        if not termo:
            self.repopular_tree_historico()
            return
        
        # Filtra por nome de arquivo que contenha o termo
        count = 0
        for item_id, dados in self.dados_historico.items():
            arquivo = dados.get('arquivo_origem', '').upper()
            
            if termo in arquivo:
                count += 1
                
                # Definir tags baseadas no status de exclusão
                tags = ["unchecked"]
                ja_excluido = dados.get('ja_excluido', False)
                if ja_excluido:
                    tags.append("erro")
                    status_display = "EXCLUÍDO"
                else:
                    tags.append("sucesso")
                    status_display = dados.get('status_envio', '')
                
                self.tree_hist.insert("", "end", iid=item_id, tags=tuple(tags),
                                     values=(count, dados.get('data_envio', ''), 
                                           dados.get('arquivo_origem', ''),
                                           dados.get('data_exclusao', ''), status_display))

    def repopular_tree_historico(self):
        """Repopula a árvore de histórico sem filtro"""
        count = 0
        for item_id, dados in self.dados_historico.items():
            count += 1
            
            # Definir tags baseadas no status de exclusão
            tags = ["unchecked"]
            ja_excluido = dados.get('ja_excluido', False)
            if ja_excluido:
                tags.append("erro")
                status_display = "EXCLUÍDO"
            else:
                tags.append("sucesso")
                status_display = dados.get('status_envio', '')
            
            self.tree_hist.insert("", "end", iid=item_id, tags=tuple(tags),
                                 values=(count, dados.get('data_envio', ''), 
                                       dados.get('arquivo_origem', ''),
                                       dados.get('data_exclusao', ''), status_display))

    def filtrar_upload(self):
        """Filtra a tabela de upload por nome de arquivo"""
        termo = self.var_busca_upload.get().strip().upper()
        
        # Limpa a tabela
        for item in self.tree_up.get_children():
            self.tree_up.delete(item)
        
        # Se não há termo, mostra todos
        if not termo:
            self.repopular_tree_upload()
            return
        
        # Filtra por nome de arquivo que contenha o termo
        for item_id, dados in self.dados_upload.items():
            arquivo = dados.get('arquivo', '').upper()
            
            if termo in arquivo:
                status = dados.get('status', 'Pronto para Envio')
                tags = ["unchecked"]
                
                if "Sucesso" in status:
                    tags = ["unchecked", "sucesso"]
                elif "ERRO" in status:
                    tags.append("erro")
                
                self.tree_up.insert("", "end", iid=item_id, tags=tuple(tags),
                                   values=(item_id, status, dados.get('arquivo', '')))

    def repopular_tree_upload(self):
        """Repopula a árvore de upload sem filtro"""
        for item_id, dados in self.dados_upload.items():
            status = dados.get('status', 'Pronto para Envio')
            tags = ["unchecked"]
            
            if "Sucesso" in status:
                tags = ["unchecked", "sucesso"]
            elif "ERRO" in status:
                tags.append("erro")
            
            self.tree_up.insert("", "end", iid=item_id, tags=tuple(tags),
                               values=(item_id, status, dados.get('arquivo', '')))

    def abrir_visualizador_historico(self, event):
        """Abre visualizador para arquivos do histórico"""
        row_id = self.tree_hist.identify_row(event.y)
        if not row_id:
            return
            
        dados = self.dados_historico.get(row_id, {})
        arquivo_nome = dados.get('arquivo_origem', '')
        ja_excluido = dados.get('ja_excluido', False)
        
        # Verificar se arquivo já foi excluído automaticamente
        if ja_excluido:
            messagebox.showinfo("Arquivo Excluído", 
                               f"O arquivo {arquivo_nome} foi excluído automaticamente.\n\n"
                               f"Arquivos são removidos 2 dias após o envio para\n"
                               f"economizar espaço em disco.")
            return
        
        # Tentar encontrar arquivo na pasta de enviados
        arquivo_path = os.path.join(DIR_ENVIADOS, arquivo_nome)
        
        if not os.path.exists(arquivo_path):
            # Se não encontrar na pasta de enviados, tentar na pasta de destino
            arquivo_path = os.path.join(DIR_DESTINO, arquivo_nome)
            
        if not os.path.exists(arquivo_path):
            messagebox.showinfo("Arquivo não encontrado", 
                               f"O arquivo {arquivo_nome} não foi encontrado.\n\n"
                               f"Local esperado: {DIR_ENVIADOS}\n"
                               f"Pode ter sido excluído automaticamente.")
            return
        
        # Criar janela de visualização
        self._criar_janela_visualizacao(arquivo_path, f"Histórico: {arquivo_nome}")

    def executar_reenvio(self):
        """Executa reenvio de arquivos selecionados do histórico"""
        jobs = [iid for iid in self.tree_hist.get_children() if "checked" in self.tree_hist.item(iid, "tags")]
        
        if not jobs:
            messagebox.showinfo("Aviso", "📋 Nenhum arquivo selecionado para reenvio.")
            return
        
        if not self.driver:
            messagebox.showerror("Erro", f"⚠️ Navegador não está pronto!\n\nStatus DHL Link: {self.dhl_status}")
            return
        
        total_selecionados = len(jobs)
        resposta = messagebox.askyesno("Confirmar Reenvio", 
            f"🔄 Reenviar {pluralizar(total_selecionados, 'arquivo')} do histórico?\n\n"
            f"Os arquivos serão reenviados para o DHL Link.\n\n"
            f"Continuar?")
        
        if not resposta:
            return
        
        # Aqui seria implementada a lógica de reenvio
        messagebox.showinfo("Funcionalidade em Desenvolvimento", 
                           "A funcionalidade de reenvio será implementada em breve.")

    def _criar_janela_visualizacao(self, arquivo_path, titulo):
        """Cria janela de visualização modal para XML"""
        if not os.path.exists(arquivo_path):
            messagebox.showinfo("Aviso", "Arquivo não encontrado.")
            return

        try:
            with open(arquivo_path, "r", encoding="utf-16") as f:
                txt = f.read()
        except:
            try:
                with open(arquivo_path, "r", encoding="utf-8") as f:
                    txt = f.read()
            except:
                txt = "Erro de leitura."

        win = tk.Toplevel(self.root)
        win.title(titulo)
        win.geometry("900x600")
        win.geometry(f"+{(win.winfo_screenwidth() // 2) - (450)}+{(win.winfo_screenheight() // 2) - (300)}")
        win.configure(bg=BG_WHITE)
        win.transient(self.root)  # Modal
        win.grab_set()  # Não permite clicar fora
        self.configurar_icone_janela(win)
        self.configurar_tecla_esc(win)  # Permite fechar com ESC

        tk.Label(win, text=f"VISUALIZADOR: {titulo}", bg=DHL_YELLOW, fg=DHL_RED, 
                font=("Segoe UI", 12, "bold")).pack(fill=tk.X, pady=(0, 10), ipady=10)

        st = scrolledtext.ScrolledText(win, font=("Consolas", 9), bg="#F9F9F9")
        st.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        st.insert(tk.END, txt)
        st.config(state="disabled")

        def copiar():
            win.clipboard_clear()
            win.clipboard_append(txt)
            messagebox.showinfo("Copiado", "Conteúdo copiado!", parent=win)

        btn_frame = tk.Frame(win, bg=BG_WHITE)
        btn_frame.pack(fill=tk.X, pady=10)
        
        tk.Button(btn_frame, text="📋 Copiar XML", command=copiar, bg=DHL_RED, fg="white", 
                 font=("Segoe UI", 10, "bold"), pady=5).pack(side=tk.LEFT, padx=20)
        
        tk.Button(btn_frame, text="❌ Fechar", command=win.destroy, bg="#95A5A6", fg="white",
                 font=("Segoe UI", 10, "bold"), pady=5).pack(side=tk.RIGHT, padx=20)

    def extrair_dados_nfe(self, caminho_arquivo):
        try:
            tree = ET.parse(caminho_arquivo)
            root = tree.getroot()
            ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
            infNFe = root.find('.//nfe:infNFe', ns)
            chave = infNFe.attrib['Id'][3:] if (
                        infNFe is not None and 'Id' in infNFe.attrib) else "CHAVE_NAO_ENCONTRADA"
            nnf = root.find('.//nfe:nNF', ns).text.zfill(9) if root.find('.//nfe:nNF', ns) is not None else "000000000"
            serie = root.find('.//nfe:serie', ns).text if root.find('.//nfe:serie', ns) is not None else "0"
            pedido = "PEDIDO_NAO_ENCONTRADO"
            infcpl = root.find('.//nfe:infCpl', ns)
            if infcpl is not None and infcpl.text:
                match = re.search(r'ORD-[A-Za-z0-9\-]+', infcpl.text)
                if match: pedido = match.group(0)
            return chave, nnf, serie, pedido
        except Exception as e:
            return None, None, None, str(e)

    def executar_conversao(self):
        jobs = [iid for iid in self.tree_conv.get_children() if
                "checked" in self.tree_conv.item(iid, "tags") and self.dados_conversao[iid]['status'] != "Convertido"]
        if not jobs: 
            return messagebox.showinfo("Aviso", "📋 Nenhum arquivo pendente selecionado para conversão.")

        # Confirmação antes de iniciar conversão
        total_selecionados = len(jobs)
        resposta = messagebox.askyesno("Confirmar Conversão", 
            f"🔄 Iniciar conversão de {total_selecionados} XML(s) para Base64?\n\n"
            f"⚠️ Os arquivos originais serão removidos da pasta após conversão.\n"
            f"✅ Arquivos convertidos serão salvos na pasta de destino.\n\n"
            f"Continuar?")
        
        if not resposta:
            return

        # Desabilita botão (Canvas)
        self.btn_acao_conv.configure(state='disabled')
        self.btn_acao_conv.unbind("<Button-1>")
        threading.Thread(target=self.thread_conversao, args=(jobs,)).start()

    def thread_conversao(self, jobs):
        sucessos, erros = 0, 0
        total_jobs = len(jobs)
        self.update_status(f"Convertendo XMLs: 0 de {total_jobs} processados...")

        for i, item_id in enumerate(jobs):
            start_time = time.time()  # Início do tempo de processamento
            job = self.dados_conversao[item_id]
            cam_orig = job['path_origem']
            chave, nnf, serie, pedido = self.extrair_dados_nfe(cam_orig)
            if not chave or chave == "CHAVE_NAO_ENCONTRADA": chave = job['original'].split('.')[0]

            try:
                with open(cam_orig, "rb") as f:
                    b64_str = base64.encodebytes(f.read()).decode('ascii').strip()

                xml_out = f"""<?xml version="1.0" encoding="utf-16"?>
<NFBASE64 xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema">
    <OrderId>{pedido}</OrderId>
    <Base64Content>\n{b64_str}\n    </Base64Content>
    <NotaFiscalNumber>{nnf}</NotaFiscalNumber>
    <NotaFiscalSerie>{serie}</NotaFiscalSerie>
</NFBASE64>"""

                nome_out = f"NFXML_BR0194SPX_{chave}.xml"
                cam_out = os.path.join(DIR_DESTINO, nome_out)

                with open(cam_out, "w", encoding="utf-16") as f_out:
                    f_out.write(xml_out)

                try:
                    os.remove(cam_orig)
                except Exception:
                    pass

                job['status'] = "Convertido";
                job['path_convertido'] = cam_out
                self.tree_conv.set(item_id, "chave", chave[:20] + "..." if len(chave) > 20 else chave)
                self.tree_conv.set(item_id, "status", "Convertido")
                self.tree_conv.set(item_id, "xml_convertido", nome_out)
                t = list(self.tree_conv.item(item_id, "tags"));
                t.remove("checked");
                t.append("unchecked");
                t.append("concluido")
                self.tree_conv.item(item_id, tags=tuple(t))
                sucessos += 1
                
                # Log detalhado da conversão
                AuditLogger.log("CONVERSÃO", f"XML convertido com sucesso - Pedido: {pedido}, NFe: {nnf}",
                               arquivo_origem=job['original'], arquivo_destino=nome_out, 
                               status_final="SUCESSO", tempo_processamento=f"{time.time() - start_time:.2f}s")
                
                # Atualizar status de progresso em tempo real
                processados = i + 1
                self.update_status(f"Convertendo XMLs: {processados} de {total_jobs} processados (Sucessos: {sucessos}, Erros: {erros})")
                
            except Exception as e:
                self.tree_conv.set(item_id, "status", "ERRO")
                t = list(self.tree_conv.item(item_id, "tags"));
                t.append("erro")
                self.tree_conv.item(item_id, tags=tuple(t))
                erros += 1
                
                # Log detalhado do erro
                AuditLogger.log("CONVERSÃO_ERRO", f"Falha na conversão: {str(e)}",
                               arquivo_origem=job['original'], arquivo_destino="", 
                               status_final="ERRO", tempo_processamento=f"{time.time() - start_time:.2f}s")
                
                # Atualizar status de progresso em tempo real
                processados = i + 1
                self.update_status(f"Convertendo XMLs: {processados} de {total_jobs} processados (Sucessos: {sucessos}, Erros: {erros})")

        # Reabilita botão (Canvas)
        try:
            self.btn_acao_conv.configure(state='normal')
            self.btn_acao_conv.bind("<Button-1>", lambda e: self.executar_conversao())
        except:
            pass
        self.calcular_resumo(self.tree_conv, self.lbl_resumo_conv, self.btn_toggle_conv)
        self.update_status(f"Conversão concluída! Total: {total_jobs} | Sucessos: {sucessos} | Erros: {erros}")
        
        # Atualizar status da aba
        if hasattr(self, 'lbl_status_conv'):
            # Define as palavras baseadas nas quantidades
            str_sucesso = "sucesso" if sucessos == 1 else "sucessos"
            str_erro = "erro" if erros == 1 else "erros"
            str_arquivo = "arquivo convertido" if sucessos == 1 else "arquivos convertidos"
            
            if erros > 0:
                self.lbl_status_conv.config(text=f" ✅ Conversão concluída: {sucessos} {str_sucesso}, {erros} {str_erro}")
            else:
                self.lbl_status_conv.config(text=f" ✅ Conversão concluída: {sucessos} {str_arquivo} com {str_sucesso}")
                
        # Variáveis com letra maiúscula para o Log
        str_conv_log = "Convertido" if sucessos == 1 else "Convertidos"
        str_erro_log = "Erro" if erros == 1 else "Erros"

        AuditLogger.log("CONVERSÃO", f"Finalizado. {sucessos} {str_conv_log}, {erros} {str_erro_log}.")

    # ==========================================================================
    # --- LÓGICA: ABA 2 (UPLOAD DHL LINK - SELENIUM) ---
    # ==========================================================================
    def iniciar_scan_upload(self):
        self.update_status("🔍 Buscando XMLs convertidos na pasta de destino...")
        for item in self.tree_up.get_children(): self.tree_up.delete(item)
        self.dados_upload.clear()

        arquivos = [f for f in os.listdir(DIR_DESTINO) if f.lower().endswith('.xml')]
        for i, f in enumerate(arquivos, 1):
            item_id = str(i)
            self.dados_upload[item_id] = {'arquivo': f, 'path': os.path.join(DIR_DESTINO, f),
                                          'status': "Pronto para Envio"}
            self.tree_up.insert("", "end", iid=item_id, tags=("checked", "even_row" if i % 2 == 0 else "odd_row"),
                                values=(i, "Pronto para Envio", f))

        self.calcular_resumo(self.tree_up, self.lbl_resumo_up, self.btn_toggle_up)
        total_arquivos = len(arquivos)
        if total_arquivos > 0:
            self.update_status(f"📤 {total_arquivos} XMLs prontos para upload | DHL Link: {self.dhl_status}")
            if hasattr(self, 'lbl_status_upload'):
                if "✅" in self.dhl_status:
                    self.lbl_status_upload.config(text=f" ✅ {total_arquivos} XMLs prontos - DHL Link conectado")
                else:
                    self.lbl_status_upload.config(text=f" ⏳ {total_arquivos} XMLs prontos - aguardando DHL Link")
        else:
            self.update_status("ℹ️ Nenhum XML convertido encontrado para upload")
            if hasattr(self, 'lbl_status_upload'):
                self.lbl_status_upload.config(text="ℹ️ Nenhum XML convertido encontrado")

    def executar_upload(self):
        if not self.driver: 
            return messagebox.showerror("Erro", f"Navegador não está pronto!\n\nStatus DHL Link: {self.dhl_status}\nO navegador Edge ainda está sendo carregado em segundo plano.\nAguarde alguns instantes e tente novamente.")

        jobs = [iid for iid in self.tree_up.get_children() if
                "checked" in self.tree_up.item(iid, "tags") and "Sucesso" not in self.dados_upload[iid]['status']]
        if not jobs: 
            return messagebox.showinfo("Aviso", "Nenhum arquivo pendente selecionado para envio.")

        # Confirmação antes de iniciar upload
        total_selecionados = len(jobs)
        lotes_necessarios = (total_selecionados + 99) // 100  # Ceiling division
        
        resposta = messagebox.askyesno("Confirmar Upload", 
            f"🚀 Iniciar upload de {total_selecionados} arquivos para DHL Link?\n\n"
            f"📦 Processamento: {lotes_necessarios} lote(s) de 100 arquivos cada\n"
            f"Continuar com o envio?")
        
        if not resposta:
            return
        
        # Desabilita botão (Canvas)
        self.btn_acao_up.configure(state='disabled')
        self.btn_acao_up.unbind("<Button-1>")
        
        # Reset da página DHL Link antes do upload
        if self.driver:
            try:
                self.update_status("Recarregando a página DHL Link...")
                self.driver.get(URL_DHL_LINK)
                time.sleep(3)
                
                # Campos não precisam ser preenchidos novamente (usam histórico do navegador)
            except:
                pass
        
        self.update_status(f"Iniciando envio de {total_selecionados} arquivos em {lotes_necessarios} lote(s)...")
        threading.Thread(target=self.thread_upload, args=(jobs,)).start()

    def thread_upload(self, jobs):
        try:
            # Dividindo em LOTES DE 100
            tamanho_lote = 100
            lotes = [jobs[i:i + tamanho_lote] for i in range(0, len(jobs), tamanho_lote)]
            total_arquivos = len(jobs)
            processados = 0

            AuditLogger.log("UPLOAD", f"Iniciado. Lotes: {len(lotes)}. Total de arquivos: {total_arquivos}")

            for num_lote, lote in enumerate(lotes, 1):
                self.update_status(f"Enviando Lote {num_lote}/{len(lotes)} - {len(lote)} arquivos (Processados: {processados}/{total_arquivos})...")

                # GARANTIA: Dá um refresh na página de Upload antes de iniciar um novo lote
                # Isso limpa a tela de sucesso/erro anterior do site
                if num_lote > 1:
                    self.driver.get(URL_DHL_LINK)
                    time.sleep(3)

                caminhos_lote = [self.dados_upload[iid]['path'] for iid in lote]
                string_caminhos = "\n".join(caminhos_lote)  # Separa por Enter

                try:
                    wait = WebDriverWait(self.driver, 30)

                    # 1. Encontrar o Input de Arquivo
                    campo_arquivo = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='file']")))

                    # 2. Injetar todos os 100 caminhos de uma vez
                    campo_arquivo.send_keys(string_caminhos)
                    time.sleep(1.5)  # Pausa pro React/Angular da DHL registrar as modificações

                    # 3. Encontrar e clicar no Botão Upload
                    botao_upload = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Upload')]")))
                    self.driver.execute_script("arguments[0].click();", botao_upload)

                    # 4. Aguardar o processamento dos 100 arquivos no site
                    # 15 Segundos para o lote de 100 fazer upload no servidor da DHL
                    time.sleep(15)

                    # 5. Ler o código HTML atual para checar as mensagens do sistema da DHL
                    html_fonte = self.driver.page_source.lower()

                    for item_id in lote:
                        nome_arq = self.dados_upload[item_id]['arquivo'].lower()

                        # VALIDAÇÃO DE ERRO: Lê se a palavra 'erro' ou 'failed' está na tela,
                        # atrelada ao nome do arquivo.
                        if f"{nome_arq} - erro" in html_fonte or "upload failed" in html_fonte:
                            self.tree_up.set(item_id, "status", "ERRO NO SITE")
                            t = list(self.tree_up.item(item_id, "tags"));
                            t.append("erro")
                            self.tree_up.item(item_id, tags=tuple(t))
                            
                            # Log detalhado do erro de upload
                            AuditLogger.log("UPLOAD_ERRO", f"Arquivo rejeitado pelo DHL Link - Lote {num_lote}",
                                           arquivo_origem=nome_arq, arquivo_destino="DHL_LINK", 
                                           status_final="REJEITADO")
                        else:
                            self.tree_up.set(item_id, "status", "Sucesso - Enviado")
                            t = list(self.tree_up.item(item_id, "tags"));
                            t.remove("checked");
                            t.append("unchecked");
                            t.append("sucesso")
                            self.tree_up.item(item_id, tags=tuple(t))

                            # Log detalhado do sucesso do upload
                            AuditLogger.log("UPLOAD", f"Arquivo enviado com sucesso - Lote {num_lote}",
                                           arquivo_origem=nome_arq, arquivo_destino="DHL_LINK", 
                                           status_final="ENVIADO")

                            # Arquivo enviado com sucesso é movido para pasta de enviados
                            try:
                                caminho_atual = self.dados_upload[item_id]['path']
                                caminho_enviados = os.path.join(DIR_ENVIADOS, nome_arq)
                                shutil.move(caminho_atual, caminho_enviados)
                                AuditLogger.log("ARQUIVO_MOVIDO", f"Movido para pasta de enviados: {nome_arq}",
                                               arquivo_origem=caminho_atual, arquivo_destino=caminho_enviados, 
                                               status_final="MOVIDO")
                            except Exception as e:
                                AuditLogger.log("ERRO_MOVER", f"Erro ao mover arquivo: {str(e)}",
                                               arquivo_origem=nome_arq, status_final="ERRO")

                except Exception as ex_lote:
                    for item_id in lote:
                        self.tree_up.set(item_id, "status", "FALHA DE COMUNICAÇÃO WEB")
                    AuditLogger.log("ERRO_SELENIUM", f"Falha no lote {num_lote}: {str(ex_lote)}")

                processados += len(lote)
                self.update_status(f"Envio: {processados} de {total_arquivos} arquivos enviados (Lote {num_lote}/{len(lotes)} concluído)")

            # Calcular estatísticas finais
            sucessos_upload = len([item for item in self.dados_upload.values() if "Sucesso" in item.get('status', '')])
            erros_upload = len([item for item in self.dados_upload.values() if "ERRO" in item.get('status', '')])
            
            self.calcular_resumo(self.tree_up, self.lbl_resumo_up, self.btn_toggle_up)
            self.update_status(f"✅ Upload concluído!")
            messagebox.showinfo("Upload Concluído", 
                f"Envio finalizado!\n\n"
                f"Total processado: {total_arquivos}\n\n"
                f"Verifique a coluna Status para detalhes.")

        except Exception as e:
            messagebox.showerror("Erro Crítico", f"Falha no processo de envio:\n{str(e)}")
        finally:
            # Reabilita botão (Canvas)
            try:
                self.btn_acao_up.configure(state='normal')
                self.btn_acao_up.bind("<Button-1>", lambda e: self.executar_upload())
            except:
                pass

    # ==========================================================================
    # --- VISUALIZADOR E LOGS ---
    # ==========================================================================
    def abrir_visualizador(self, event, tree, dict_dados):
        row_id = tree.identify_row(event.y)
        if not row_id: return
        job = dict_dados[row_id]

        path = job.get('path_convertido') or job.get('path_origem')
        if not path or not os.path.exists(path): return messagebox.showinfo("Aviso",
                                                                            "Arquivo não encontrado fisicamente.")

        try:
            with open(path, "r", encoding="utf-16") as f:
                txt = f.read()
        except:
            try:
                with open(path, "r", encoding="utf-8") as f:
                    txt = f.read()
            except:
                txt = "Erro de leitura."

        win = tk.Toplevel(self.root)
        win.title("Visualizador XML")
        win.geometry("900x600")
        win.geometry(f"+{(win.winfo_screenwidth() // 2) - (450)}+{(win.winfo_screenheight() // 2) - (300)}")
        win.configure(bg=BG_WHITE)
        win.transient(self.root)  # Modal
        win.grab_set()  # Não permite clicar fora
        self.configurar_icone_janela(win)
        self.configurar_tecla_esc(win)  # Permite fechar com ESC

        tk.Label(win, text="CONTEÚDO DO ARQUIVO", bg=DHL_YELLOW, fg=DHL_RED, font=("Segoe UI", 12, "bold")).pack(
            fill=tk.X, pady=(0, 10), ipady=10)

        st = scrolledtext.ScrolledText(win, font=("Consolas", 9), bg="#F9F9F9")
        st.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        st.insert(tk.END, txt)
        st.config(state="disabled")

        def copiar():
            self.root.clipboard_clear();
            self.root.clipboard_append(txt)
            messagebox.showinfo("Copiado", "Conteúdo copiado!", parent=win)

        tk.Button(win, text="📋 Copiar XML", command=copiar, bg=DHL_RED, fg="white", font=("Segoe UI", 10, "bold"),
                  pady=5).pack(pady=10)

    def abrir_janela_log(self):
        win = tk.Toplevel(self.root)
        win.title("Log de Auditoria - Hoje")
        win.geometry("700x500")
        win.geometry(f"+{(win.winfo_screenwidth() // 2) - (700 // 2)}+{(win.winfo_screenheight() // 2) - (500 // 2)}")

        tk.Label(win, text="REGISTROS DE ATIVIDADE (HOJE)", bg=DHL_YELLOW, fg=DHL_RED,
                 font=("Segoe UI", 12, "bold")).pack(fill=tk.X, ipady=10)

        frame = tk.Frame(win, bg=BG_WHITE);
        frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        st = scrolledtext.ScrolledText(frame, font=("Consolas", 9), bg="#F9F9F9", relief="solid")
        st.pack(fill=tk.BOTH, expand=True)
        st.insert(tk.END, AuditLogger.ler_logs_do_dia())
        st.configure(state='disabled')


if __name__ == "__main__":
    root = tk.Tk()
    app = ConversorXML_Starlink(root)
    root.mainloop()