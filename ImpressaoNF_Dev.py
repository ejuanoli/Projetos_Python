import os
import sys
import shutil
import re
import threading
import time
import datetime
import socket
import tkinter as tk
import concurrent.futures
from tkinter import ttk, messagebox
import win32print
import subprocess
from pypdf import PdfReader, PdfWriter
import webbrowser
import logging
from openpyxl import Workbook, load_workbook


CLIENTES_CONFIG = {
    "SCHNEIDER": {
        "keywords": ["SCHNEIDER", "ELECTRIC"],
        "recusa": True  # Imprime duplex com formulário atrás
    },
    "DELL": {
        "keywords": ["DELL"],
        "recusa": False # Imprime normal (simplex)
    },
    "STARLINK": {
        "keywords": ["STARLINK", "SPACE"],
        "recusa": False
    }
}

# Configuração padrão caso não ache nenhuma palavra-chave
CLIENTE_PADRAO = ""

logger = logging.getLogger("pypdf")
logger.setLevel(logging.ERROR)

class AutoScrollbar(ttk.Scrollbar):
    def set(self, lo, hi):
        if float(lo) <= 0.0 and float(hi) >= 1.0:
            self.pack_forget()
        else:
            self.pack(side=tk.RIGHT, fill=tk.Y)
        ttk.Scrollbar.set(self, lo, hi)

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

USER_HOME = os.path.expanduser("~")
#CAMINHO_RELATIVO = r"Downloads\Teste_NF"
#CAMINHO_RELATIVO = r"DPDHL\Imp_Schneider - Imp\Arquivos" #USUÁRIO FINAL
CAMINHO_RELATIVO = r"DPDHL\Imp_Schneider - Documentos\Imp\Arquivos" #DEV
BASE_DIR = os.path.join(USER_HOME, CAMINHO_RELATIVO)
URL_SHAREPOINT = "https://dpdhl.sharepoint.com/teams/Imp_Schneider/Shared%20Documents/Forms/AllItems.aspx?id=%2Fteams%2FImp%5FSchneider%2FShared%20Documents%2FImp&viewid=67e6e6c0%2Dd2ba%2D4d8e%2D903d%2D4ddcc2c018b9"

if not os.path.exists(BASE_DIR):
    def abrir_sharepoint(event):
        webbrowser.open(URL_SHAREPOINT)

    err_window = tk.Tk()
    err_window.title("Ação Necessária - Sincronização")
    largura, altura = 550, 750
    pos_x = (err_window.winfo_screenwidth() // 2) - (largura // 2)
    pos_y = (err_window.winfo_screenheight() // 2) - (altura // 2)
    err_window.geometry(f"{largura}x{altura}+{pos_x}+{pos_y}")
    err_window.configure(bg="white")

    try:
        err_window.iconbitmap(resource_path("logo.ico"))
    except:
        pass

    tk.Label(err_window, text="PASTA NÃO SINCRONIZADA",
             fg="#D40511", bg="white", font=("Segoe UI", 16, "bold")).pack(pady=(15, 5))
    msg = (f"O sistema não encontrou a pasta 'Arquivos' no seu computador.\n"
           f"Caminho esperado: {BASE_DIR}")
    tk.Label(err_window, text=msg, bg="white", font=("Segoe UI", 10)).pack(pady=5)

    try:
        img_path = resource_path("tutorial_sync.png")
        png_tutorial = tk.PhotoImage(file=img_path)
        lbl_img = tk.Label(err_window, image=png_tutorial, bg="white", bd=1, relief="solid")
        lbl_img.image = png_tutorial
        lbl_img.pack(pady=10)
    except Exception:
        tk.Label(err_window, text="[Imagem tutorial_sync.png não encontrada]", bg="white", fg="gray").pack()

    instrucoes = ("Você deve clicar nos 3 pontos na parte superior do Sharepoint "
                  "e clicar no botão Sincronizar (ou Sync).\n\n"
                  "Um popup irá surgir pedindo permissão para sincronizar. "
                  "Após isso, o programa deve ser reiniciado.")
    tk.Label(err_window, text=instrucoes, bg="white", fg="#333",
             font=("Segoe UI", 10), justify="center", wraplength=500).pack(pady=5)

    lbl_link = tk.Label(err_window, text=">>> CLIQUE AQUI PARA ABRIR O SHAREPOINT <<<",
                        fg="blue", bg="white", cursor="hand2", font=("Segoe UI", 11, "bold", "underline"))
    lbl_link.pack(pady=15)
    lbl_link.bind("<Button-1>", abrir_sharepoint)

    tk.Button(err_window, text="Fechar Programa", command=sys.exit,
              bg="#333", fg="white", font=("Segoe UI", 10), width=20).pack(pady=10)
    err_window.mainloop()
    sys.exit()

# --- PREFERÊNCIAS ---
CONFIG_FILE = os.path.join(USER_HOME, ".dhl_printer_config.txt")

# --- DEFINIÇÃO DAS SUBPASTAS ---# --- DEFINIÇÃO DAS SUBPASTAS ---
DIR_ENTRADA = os.path.join(BASE_DIR, "1_PastaOrigem")
DIR_DESTINO = os.path.join(BASE_DIR, "2_PastaDestino")
DIR_LOGS = os.path.join(BASE_DIR, "3_StatusImpressoes")
DIR_HISTORICO = os.path.join(BASE_DIR, "4_HistoricoAntigo")
SUMATRA_PATH = os.path.join(BASE_DIR, "SumatraPDF.exe")

for pasta in [DIR_ENTRADA, DIR_DESTINO, DIR_LOGS, DIR_HISTORICO]:
    if not os.path.exists(pasta):
        try:
            os.makedirs(pasta)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível criar pasta: {pasta}\n{e}")

# ==============================================================================
# --- CORES E ESTILOS ---
# ==============================================================================
DHL_YELLOW = "#FFCC00"
DHL_RED = "#D40511"
DHL_BLACK = "#000000"
BG_WHITE = "#FFFFFF"
BG_GREY = "#F2F2F2"


# ==============================================================================
# --- SISTEMA DE LOG E AUDITORIA ---
# ==============================================================================
class AuditLogger:
    @staticmethod
    def log(acao, detalhes=""):
        try:
            usuario = os.getlogin()
            maquina = socket.gethostname()
            data_hoje = datetime.datetime.now().strftime("%Y-%m-%d")

            nome_arquivo = f"Audit_{data_hoje}_{usuario}_{maquina}.txt"
            caminho_log = os.path.join(DIR_LOGS, nome_arquivo)

            timestamp = datetime.datetime.now().strftime("%H:%M:%S")
            linha = f"[{timestamp}] [{usuario}] {acao.upper()} - {detalhes}\n"

            with open(caminho_log, "a", encoding="utf-8") as f:
                f.write(linha)

        except Exception as e:
            print(f"Falha ao gravar log de auditoria: {e}")

    @staticmethod
    def ler_logs_do_dia():
        try:
            usuario = os.getlogin()
            maquina = socket.gethostname()
            data_hoje = datetime.datetime.now().strftime("%Y-%m-%d")
            nome_arquivo = f"Audit_{data_hoje}_{usuario}_{maquina}.txt"
            caminho_log = os.path.join(DIR_LOGS, nome_arquivo)

            if os.path.exists(caminho_log):
                with open(caminho_log, "r", encoding="utf-8") as f:
                    return f.read()
            return "Nenhum registro encontrado para hoje."
        except Exception:
            return "Erro ao ler arquivo de log."


# ==============================================================================
# --- CLASSE PRINCIPAL ---
# ==============================================================================
class ImpressorDHL:
    def __init__(self, root):
        self.root = root
        self.root.title("DHL Fiscal Print Manager")

        # Log de Abertura
        AuditLogger.log("SISTEMA", "Programa Iniciado")

        try:
            self.root.state('zoomed')
        except:
            self.root.attributes('-fullscreen', True)

        self.root.configure(bg=BG_WHITE)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        try:
            icon_path = resource_path("logo.ico")
            self.root.iconbitmap(icon_path)
        except:
            pass

        # === CORREÇÃO: Variáveis separadas para evitar recarregamento ===
        self.dados_entrada = {}  # Cache Aba 1
        self.dados_saida = {}  # Cache Aba 2
        self.dados_atuais = {}  # Ponteiro para o dicionário ativo no momento
        self.scan_ativo = False  # Trava para evitar erro de threads colidindo

        self._criar_icones_dinamicamente()
        self._configurar_estilos()
        self._criar_layout()
        self._carregar_preferencias()

        # Arquivamento automático
        self.root.after(100, self.realizar_arquivamento_historico)

        # Inicia varredura APENAS na aba inicial
        self.root.after(800, lambda: self.iniciar_scan(forcar_refresh=True))

    def on_close(self):
        AuditLogger.log("SISTEMA", "Programa Fechado pelo Usuário")
        self.root.destroy()

    def realizar_arquivamento_historico(self):
        """
        Move arquivos da pasta DESTINO para HISTORICO se tiverem sido impressos há mais de 24h.
        Baseia-se no Timestamp adicionado ao nome do arquivo durante a impressão.
        """
        try:
            count = 0
            agora = time.time()
            limite_tempo = 24 * 3600  # 24 horas em segundos

            arquivos = os.listdir(DIR_DESTINO)
            for f in arquivos:
                full_path = os.path.join(DIR_DESTINO, f)

                mover = False

                # Tenta ler o Timestamp do nome do arquivo (formato: 1728394___Nome.pdf)
                if "___" in f:
                    try:
                        timestamp_str = f.split("___")[0]
                        timestamp_arq = float(timestamp_str)

                        # Se a diferença entre AGORA e a IMPRESSÃO for maior que 24h
                        if (agora - timestamp_arq) > limite_tempo:
                            mover = True
                    except ValueError:
                        # Se não conseguir converter o numero, ignora
                        pass

                if mover:
                    destino_hist = os.path.join(DIR_HISTORICO, f)
                    try:
                        shutil.move(full_path, destino_hist)
                        AuditLogger.log("AUTO_ARQUIVO", f"Movido para Histórico (+24h): {f}")
                        count += 1
                    except Exception as e:
                        print(f"Erro ao arquivar {f}: {e}")

            if count > 0:
                print(f"Arquivamento: {count} arquivos movidos.")

        except Exception as e:
            AuditLogger.log("ERRO_ARQUIVAMENTO", str(e))

    def _configurar_estilos(self):
        style = ttk.Style()
        style.theme_use('clam')

        # --- TABELA (TREEVIEW) ---
        style.configure("Treeview", background=BG_WHITE, foreground="black",
                        fieldbackground=BG_WHITE, rowheight=45, font=('Segoe UI', 9))
        style.map("Treeview",
                  background=[('selected', BG_WHITE), ('active', BG_WHITE)],
                  foreground=[('selected', 'black'), ('active', 'black')])

        style.configure("Treeview.Heading", background=DHL_RED, foreground="white",
                        font=('Segoe UI', 10, 'bold'), relief="raised", borderwidth=1)
        style.map("Treeview.Heading", background=[('active', "#b3040e")])

        # --- BOTÕES ---
        style.configure("Red.TButton", font=('Segoe UI', 10, 'bold'), background=DHL_RED, foreground="white",
                        borderwidth=0, focuscolor=DHL_RED)
        style.map("Red.TButton", background=[('active', "#b3040e")])

        style.configure("White.TButton", font=('Segoe UI', 9), background=BG_WHITE, foreground="black", borderwidth=1,
                        relief="solid", focuscolor=BG_WHITE)
        style.map("White.TButton", background=[('active', "#e6e6e6")])

        # --- ABAS (CORRIGIDO PARA SOBREPOSIÇÃO CORRETA) ---

        style.configure("TNotebook", background=BG_GREY, borderwidth=0)

        style.configure("TNotebook.Tab",
                        font=('Segoe UI', 10, 'bold'),
                        padding=[20, 3],
                        background="#D0D0D0",  # Cinza escuro = fundo
                        foreground="#555555",  # Texto cinza
                        borderwidth=0)

        style.map("TNotebook.Tab",
                  padding=[("selected", (20, 4))],
                  background=[("selected", DHL_YELLOW)],  # Branco = Frente
                  foreground=[("selected", DHL_RED)]
                  )

    def _round_rectangle(self, canvas, x1, y1, x2, y2, radius=25, **kwargs):
        """Desenha retângulo com cantos arredondados no Canvas"""
        points = [x1 + radius, y1,
                  x1 + radius, y1,
                  x2 - radius, y1,
                  x2 - radius, y1,
                  x2, y1,
                  x2, y1 + radius,
                  x2, y1 + radius,
                  x2, y2 - radius,
                  x2, y2 - radius,
                  x2, y2,
                  x2 - radius, y2,
                  x2 - radius, y2,
                  x1 + radius, y2,
                  x1 + radius, y2,
                  x1, y2,
                  x1, y2 - radius,
                  x1, y2 - radius,
                  x1, y1 + radius,
                  x1, y1 + radius,
                  x1, y1]
        return canvas.create_polygon(points, **kwargs, smooth=True)

    def _criar_layout(self):

        # ======================================================================
        # 1) RODAPÉ – container dedicado p/ garantir ordem sempre fixa
        # ======================================================================
        self.footer = tk.Frame(self.root, bg=DHL_YELLOW)
        self.footer.pack(side=tk.BOTTOM, fill=tk.X)

        # 1.1) Faixa preta (status) – sempre colada no chão do footer
        self.lbl_status = tk.Label(
            self.footer, text=" Sistema pronto.", bg="#333", fg="white",
            anchor="w", font=("Arial", 8)
        )
        self.lbl_status.pack(side=tk.TOP, fill=tk.X)

        # 1.2) Faixa amarela (botões) – fica acima da faixa preta
        bot_frame = tk.Frame(self.footer, bg=DHL_YELLOW, height=80)

        bot_frame.pack(side=tk.BOTTOM, fill=tk.X)
        bot_frame.pack_propagate(False)

        # Conteúdo do Rodapé Amarelo (inalterado)
        btn_scan = ttk.Button(bot_frame, text="ATUALIZAR LISTA",
                              command=lambda: self.iniciar_scan(forcar_refresh=True),
                              style="White.TButton")
        btn_scan.pack(side=tk.LEFT, padx=20, pady=15, ipadx=10, ipady=5)

        self.btn_toggle = ttk.Button(bot_frame, text="MARCAR TODOS", width=20,
                                     command=self.toggle_all, style="White.TButton")
        self.btn_toggle.pack(side=tk.LEFT, padx=5, pady=15, ipadx=10, ipady=5)

        self.lbl_resumo = tk.Label(bot_frame, text="---", bg=DHL_YELLOW, fg=DHL_RED,
                                   font=('Segoe UI', 14, 'bold'))
        self.lbl_resumo.pack(side=tk.LEFT, padx=50, pady=15)

        self.btn_acao = ttk.Button(bot_frame, text="IMPRIMIR AGORA",
                                   command=self.executar_acao_principal,
                                   style="Red.TButton")
        self.btn_acao.pack(side=tk.RIGHT, padx=20, pady=15, ipadx=20, ipady=5)

        # ======================================================================
        # 2) TOPO (cabeçalho) – como você já tinha
        # ======================================================================
        header_frame = tk.Frame(self.root, bg=DHL_YELLOW, height=100)
        header_frame.pack(side=tk.TOP, fill=tk.X)

        top_content = tk.Frame(header_frame, bg=DHL_YELLOW)
        top_content.pack(side=tk.LEFT, fill=tk.BOTH, padx=20, pady=10)

        # Logo / Título
        try:
            img_path = resource_path("dhl_logo.png")
            self.logo_image = tk.PhotoImage(file=img_path).subsample(2, 2)
            lbl_logo = tk.Label(top_content, image=self.logo_image, bg=DHL_YELLOW)
            lbl_logo.pack(side=tk.LEFT, padx=(0, 20))
        except:
            lbl_logo = tk.Label(top_content, text="DHL", bg=DHL_YELLOW, fg=DHL_RED, font=('Arial Black', 28, 'italic'))
            lbl_logo.pack(side=tk.LEFT, padx=(0, 20))

        lbl_title = tk.Label(top_content, text="GERENCIADOR DE IMPRESSÃO FISCAL", bg=DHL_YELLOW, fg=DHL_RED,
                             font=('Segoe UI', 18, 'bold', 'italic'))
        lbl_title.pack(side=tk.LEFT, anchor='center')

        # Ícones Topo
        btn_help = tk.Label(header_frame, image=self.icon_help, bg=DHL_YELLOW, cursor="hand2")
        btn_help.pack(side=tk.RIGHT, padx=(5, 20), pady=22)
        btn_help.bind("<Button-1>", lambda e: self.abrir_janela_ajuda())

        btn_log = tk.Label(header_frame, image=self.icon_log, bg=DHL_YELLOW, cursor="hand2")
        btn_log.pack(side=tk.RIGHT, padx=(20, 5), pady=20)
        btn_log.bind("<Button-1>", lambda e: self.abrir_janela_log())

        # ======================================================================
        # 3) CONTROLES (logo abaixo do topo) – como você já tinha
        # ======================================================================

        ctrl_frame = tk.Frame(self.root, bg=BG_GREY, pady=15)
        ctrl_frame.pack(side=tk.TOP, fill=tk.X)

        # Impressora (ESQUERDA)
        frame_print = tk.Frame(ctrl_frame, bg=BG_GREY)
        frame_print.pack(side=tk.LEFT, padx=40, anchor="n")

        tk.Label(frame_print, text="IMPRESSORA:", bg=BG_GREY, fg='#666',
                 font=('Segoe UI', 8, 'bold')).pack(anchor='w', pady=(0, 2))

        try:
            self.printers = [p[2] for p in win32print.EnumPrinters(
                win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)]
        except:
            self.printers = []

        self.combo_printer = ttk.Combobox(frame_print, values=self.printers, width=35,
                                          state="readonly", font=('Segoe UI', 10))
        if self.printers:
            self.combo_printer.current(0)

        self.combo_printer.pack(anchor="w", ipady=5)

        # Busca (AO LADO, também à esquerda)
        frame_busca_container = tk.Frame(ctrl_frame, bg=BG_GREY)
        frame_busca_container.pack(side=tk.LEFT, padx=40, anchor="n")

        tk.Label(frame_busca_container, text="BUSCAR Nº NOTA FISCAL:", bg=BG_GREY, fg='#666',
                 font=('Segoe UI', 8, 'bold')).pack(anchor='w', pady=(0, 2))

        self.canvas_search = tk.Canvas(frame_busca_container, width=312, height=1,
                                       bg=BG_GREY, highlightthickness=0)
        self.canvas_search.pack(anchor="w")

        # StringVar precisa existir antes do desenho do search
        self.var_busca = tk.StringVar()
        self.var_busca.trace("w", self.filtrar_tabela)

        # Agora sincroniza com a altura real do combobox
        self.root.after(0, self._sync_altura_search_combo)

        self.var_busca = tk.StringVar()
        self.var_busca.trace("w", self.filtrar_tabela)
        entry_busca = tk.Entry(self.canvas_search, textvariable=self.var_busca, width=28,
                               font=('Segoe UI', 10), bg="white", bd=0, highlightthickness=0)
        self.canvas_search.create_window(32, 18, window=entry_busca, anchor="w")

        lbl_clear = tk.Label(self.canvas_search, text="✕", bg="white", fg="#BBB",
                             font=("Arial", 9, "bold"), cursor="hand2")
        lbl_clear.bind("<Button-1>", lambda e: [self.var_busca.set(""), self.root.focus()])
        self.canvas_search.create_window(292, 18, window=lbl_clear)
        # ======================================================================
        # 4) ÁREA CENTRAL EXPANSÍVEL – notebook dentro de um container próprio
        # ======================================================================
        center = tk.Frame(self.root, bg=BG_WHITE)
        center.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        # Coloque o notebook dentro do 'center' para ele usar só o espaço central
        self.notebook = ttk.Notebook(center)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        self.tab_entrada = tk.Frame(self.notebook, bg=BG_WHITE)
        self.notebook.add(self.tab_entrada, text="  NOVOS ARQUIVOS (A IMPRIMIR)  ")

        self.tab_saida = tk.Frame(self.notebook, bg=BG_WHITE)
        self.notebook.add(self.tab_saida, text="  IMPRESSÕES REALIZADAS (REIMPRIMIR)  ")

        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_change)

        self.tree_entrada = self._criar_treeview(self.tab_entrada)
        self.tree_saida = self._criar_treeview(self.tab_saida)

        self.tree_atual = self.tree_entrada
        self.dados_atuais = self.dados_entrada

    def _sync_altura_search_combo(self):
        """Sincroniza a altura do Search (canvas) com a altura real do Combobox."""
        self.root.update_idletasks()

        combo_h = self.combo_printer.winfo_height()
        if combo_h <= 1:
            # Se ainda não mediu (raro), tenta de novo
            self.root.after(30, self._sync_altura_search_combo)
            return

        # Desenha o search com a MESMA altura do combobox
        self._desenhar_search(combo_h)

    def _desenhar_search(self, h):
        """Desenha o widget de busca no canvas com altura h (px)."""
        self.canvas_search.delete("all")
        self.canvas_search.config(height=h)

        # Centro vertical do canvas
        cy = h // 2

        # Retângulos arredondados (ajuste baseado na altura)
        # Margens para não cortar borda
        top = 2
        bottom = h - 2

        self._round_rectangle(self.canvas_search, 3, top, 310, bottom, radius=6,
                              fill="#D0D0D0", outline="")
        self._round_rectangle(self.canvas_search, 1, top - 1, 308, bottom - 1, radius=6,
                              fill="white", outline="#ACACAC")

        # Ícone lupa
        try:
            caminho_lupa = resource_path("search.png")
            img_raw = tk.PhotoImage(file=caminho_lupa)

            target_size = 18
            w = img_raw.width()
            factor = max(1, w // target_size)
            self.icon_search_png = img_raw.subsample(factor, factor)

            self.btn_lupa_id = self.canvas_search.create_image(15, cy, image=self.icon_search_png, anchor="center")

            def _cursor_mao(_):
                self.canvas_search.config(cursor="hand2")  # cursor de clique

            def _cursor_normal(_):
                self.canvas_search.config(cursor="")  # volta ao padrão (seta)

            self.canvas_search.tag_bind(self.btn_lupa_id, "<Enter>", _cursor_mao)
            self.canvas_search.tag_bind(self.btn_lupa_id, "<Leave>", _cursor_normal)

            def on_lupa_press(event):
                self.canvas_search.move(self.btn_lupa_id, 1, 1)

            def on_lupa_release(event):
                self.canvas_search.move(self.btn_lupa_id, -1, -1)
                self.root.focus()
                self.filtrar_tabela()

            self.canvas_search.tag_bind(self.btn_lupa_id, "<ButtonPress-1>", on_lupa_press)
            self.canvas_search.tag_bind(self.btn_lupa_id, "<ButtonRelease-1>", on_lupa_release)

        except Exception:
            self.canvas_search.create_text(15, cy, text="🔍", font=("Segoe UI", 10), fill="#555", anchor="center")

        # Entry + Clear
        if not hasattr(self, "var_busca"):
            self.var_busca = tk.StringVar()
            self.var_busca.trace("w", self.filtrar_tabela)

        entry_busca = tk.Entry(self.canvas_search, textvariable=self.var_busca, width=28,
                               font=('Segoe UI', 10), bg="white", bd=0, highlightthickness=0)

        self.canvas_search.create_window(32, cy, window=entry_busca, anchor="w")

        lbl_clear = tk.Label(self.canvas_search, text="✕", bg="white", fg="#BBB",
                             font=("Arial", 9, "bold"), cursor="hand2")
        lbl_clear.bind("<Button-1>", lambda e: [self.var_busca.set(""), self.root.focus()])

        # Hover (efeito visual)
        def _x_enter(_):
            lbl_clear.config(fg="#666")  # fica mais escuro ao passar o mouse

        def _x_leave(_):
            lbl_clear.config(fg="#BBB")  # volta ao normal

        lbl_clear.bind("<Enter>", _x_enter)
        lbl_clear.bind("<Leave>", _x_leave)

        self.canvas_search.create_window(292, cy, window=lbl_clear)

    def _criar_treeview(self, parent_frame):
        # 1. Definir as colunas (Adicionado 'cli' para Cliente)
        cols = ("id", "cli", "num", "chave", "nf", "cte")

        # 2. Criar Treeview
        tree = ttk.Treeview(parent_frame, columns=cols, show='tree headings', selectmode="none")

        # 3. Configurar Cabeçalhos e Colunas
        tree.heading("#0", text="SEL")
        tree.column("#0", width=40, stretch=False, anchor='center')

        tree.heading("id", text="#")
        tree.column("id", width=35, stretch=False, anchor='center')

        # --- NOVA COLUNA CLIENTE ---
        tree.heading("cli", text="CLIENTE")
        tree.column("cli", width=100, stretch=False, anchor='center')
        # ---------------------------

        tree.heading("num", text="Nº DOCUMENTO")
        tree.column("num", width=120, stretch=False, anchor='center')

        tree.heading("chave", text="CHAVES DE ACESSO")
        tree.column("chave", width=180, stretch=True, anchor='center')

        tree.heading("nf", text="ARQUIVO NF")
        tree.column("nf", width=180, stretch=True, anchor='center')

        tree.heading("cte", text="ARQUIVO CTE")
        tree.column("cte", width=180, stretch=True, anchor='center')

        # 4. Configurar Tags de Cores e Estilos
        tree.tag_configure("odd_row", background="#F2F2F2")
        tree.tag_configure("even_row", background="#FFFFFF")

        # Tag específica para destacar Schneider (opcional, visual agradável)
        tree.tag_configure("schneider_row", foreground="#000000")  # Verde escuro no texto

        tree.tag_configure("checked", image=self.img_checked)
        tree.tag_configure("unchecked", image=self.img_unchecked)
        tree.tag_configure("disabled", image=self.img_disabled, foreground="#999")

        # Scrollbar
        scrollbar = AutoScrollbar(parent_frame, orient="vertical", command=tree.yview)
        tree.configure(yscroll=scrollbar.set)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree.bind('<Button-1>', self.on_click_tree)

        return tree

    def _criar_icones_dinamicamente(self):
        size = 20
        self.img_unchecked = tk.PhotoImage(width=size, height=size)
        self.img_checked = tk.PhotoImage(width=size, height=size)
        self.img_disabled = tk.PhotoImage(width=size, height=size)

        row_data = "{" + " ".join(["#FFFFFF"] * size) + "} "
        full_data = row_data * size
        self.img_unchecked.put(full_data)
        self.img_checked.put(full_data)

        row_grey = "{" + " ".join(["#E0E0E0"] * size) + "} "
        full_grey = row_grey * size
        self.img_disabled.put(full_grey)

        # Bordas
        for i in range(size):
            for img in [self.img_unchecked, self.img_checked, self.img_disabled]:
                img.put("#888", (i, 0))
                img.put("#888", (i, size - 1))
                img.put("#888", (0, i))
                img.put("#888", (size - 1, i))

        # X Vermelho no Disabled (Visual de bloqueio)
        for i in range(4, 16):
            self.img_disabled.put("#AA0000", (i, i))
            self.img_disabled.put("#AA0000", (i, 19 - i))

        # Check Verde
        check_color = "#008000"
        for i in range(4):
            x, y = 5 + i, 11 + i
            self.img_checked.put(check_color, (x, y))
            self.img_checked.put(check_color, (x, y + 1))
        for i in range(8):
            x, y = 8 + i, 14 - i
            self.img_checked.put(check_color, (x, y))
            self.img_checked.put(check_color, (x, y + 1))
            self.img_checked.put(check_color, (x - 1, y + 1))

        self.icon_log = tk.PhotoImage(width=32, height=32)
        bg_yellow = "{" + " ".join([DHL_YELLOW] * 32) + "} "
        self.icon_log.put(bg_yellow * 32)

        white_block = "{" + " ".join(["#FFFFFF"] * 20) + "} "
        for y in range(4, 28):
            self.icon_log.put(white_block, (6, y))
        black_line = "{" + " ".join(["#000000"] * 14) + "} "
        for y in [8, 12, 16, 20, 24]:
            self.icon_log.put(black_line, (9, y))

        # --- ÍCONE DE AJUDA (Carregando help.png) ---
            try:
                path_help = resource_path("help.png")
                # 1. Carrega a imagem original em memória
                raw_img = tk.PhotoImage(file=path_help)

                target_size = 28

                w = raw_img.width()

                factor = w // target_size

                # Aplica a redução apenas se a imagem for maior que o alvo
                if factor > 1:
                    self.icon_help = raw_img.subsample(factor, factor)
                else:
                    self.icon_help = raw_img  # Usa original se já for pequena

            except Exception:
                # Fallback simples caso a imagem não exista (quadrado azul)
                self.icon_help = tk.PhotoImage(width=28, height=28)
                row = "{" + " ".join(["#0055A5"] * 28) + "}"
                data = " ".join([row] * 28)
                self.icon_help.put(data)

    def _carregar_preferencias(self):
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    ultima_impressora = f.read().strip()
                if ultima_impressora in self.printers:
                    self.combo_printer.set(ultima_impressora)
        except Exception:
            pass

    def _salvar_preferencias(self):
        try:
            impressora_atual = self.combo_printer.get()
            if impressora_atual:
                with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                    f.write(impressora_atual)
        except Exception:
            pass

    def update_status(self, msg):
        self.lbl_status.config(text=f" {msg}")
        self.root.update_idletasks()

    # --- LÓGICA DE ABAS (CORRIGIDA) ---
    def on_tab_change(self, event):
        idx = self.notebook.index("current")

        if idx == 0:
            self.tree_atual = self.tree_entrada
            self.dados_atuais = self.dados_entrada
            self.btn_acao.config(text="IMPRIMIR AGORA")
            # Se a lista estiver vazia, carrega pela primeira vez
            if not self.tree_entrada.get_children():
                self.iniciar_scan(forcar_refresh=True)
        else:
            self.tree_atual = self.tree_saida
            self.dados_atuais = self.dados_saida
            self.btn_acao.config(text="REIMPRIMIR SELEÇÃO")
            # Se a lista estiver vazia, carrega pela primeira vez
            if not self.tree_saida.get_children():
                self.iniciar_scan(forcar_refresh=True)
        self.calcular_resumo()

    def calcular_resumo(self):
        total_conjuntos = 0
        total_paginas = 0

        # Variáveis para controle do botão
        total_habilitados = 0
        total_marcados = 0

        items = self.tree_atual.get_children()

        for item_id in items:
            tags = self.tree_atual.item(item_id, "tags")

            # Se não está bloqueado, entra na conta do botão
            if "disabled" not in tags:
                total_habilitados += 1
                if "checked" in tags:
                    total_marcados += 1

            # Lógica de Soma de Páginas (apenas marcados contam)
            if "checked" in tags:
                total_conjuntos += 1
                if item_id in self.dados_atuais:
                    job = self.dados_atuais[item_id]
                    total_paginas += job['total_pgs']

        self.lbl_resumo.config(text=f"CONJUNTOS SELECIONADOS: {total_conjuntos}  |  TOTAL FOLHAS: {total_paginas}")

        # Se existem itens habilitados e TODOS eles estão marcados -> Botão vira Desmarcar
        if total_habilitados > 0 and total_habilitados == total_marcados:
            self.btn_toggle.config(text="DESMARCAR TODOS")
        else:
            self.btn_toggle.config(text="MARCAR TODOS")

    def on_click_tree(self, event):
        region = self.tree_atual.identify("region", event.x, event.y)
        if region == "heading": return

        row_id = self.tree_atual.identify_row(event.y)
        if not row_id: return

        # Obtém a coluna clicada (ex: "#1", "#6")
        col_id_str = self.tree_atual.identify_column(event.x)
        # Converte "#6" para inteiro 6
        col_idx = int(col_id_str.replace("#", ""))

        # A coluna AÇÃO é a 6ª coluna no nosso setup
        # (id, num, chave, nf, cte, acao) -> 1, 2, 3, 4, 5, 6

        if col_idx == 6:
            # Lógica de Exclusão
            if row_id in self.dados_atuais:
                job = self.dados_atuais[row_id]
                if job.get('tipo') == 'DUPLICATA':

                    caminho = job.get('caminho_completo')
                    nome = os.path.basename(caminho)

                    if messagebox.askyesno("Excluir", f"Confirma exclusão da duplicata?\n{nome}"):
                        try:
                            os.remove(caminho)
                            self.tree_atual.delete(row_id)
                            del self.dados_atuais[row_id]
                            # Feedback visual simples
                            print(f"Excluído: {nome}")
                        except Exception as e:
                            messagebox.showerror("Erro", str(e))
                    return  # Encerra aqui, não marca checkbox

        # Lógica Checkbox (Padrão)
        tags = self.tree_atual.item(row_id, "tags")
        if "duplicate_row" in tags or "disabled" in tags:
            return

        novo = "unchecked" if "checked" in tags else "checked"

        # Mantém a cor de fundo (odd/even)
        bg = "odd_row" if "odd_row" in tags else "even_row"
        self.tree_atual.item(row_id, tags=(novo, bg))

        self.calcular_resumo()

    def toggle_all(self):
        items = self.tree_atual.get_children()
        if not items: return

        texto_atual = self.btn_toggle.cget("text")
        novo_status = "checked" if texto_atual == "MARCAR TODOS" else "unchecked"

        for item_id in items:
            tags_atuais = self.tree_atual.item(item_id, "tags")

            # Pula os bloqueados
            if "disabled" in tags_atuais:
                continue

            # Mesma lógica matemática baseada no ID
            try:
                num_id = int(item_id)
                cor_fixa = "odd_row" if num_id % 2 != 0 else "even_row"
            except:
                cor_fixa = "even_row"

            self.tree_atual.item(item_id, tags=(novo_status, cor_fixa))

        self.calcular_resumo()

    def iniciar_scan(self, forcar_refresh=False):
        # Evita múltiplas threads e colisão
        if self.scan_ativo:
            print("Scan já está em andamento. Aguarde.")
            return

        idx = self.notebook.index("current")
        pasta_alvo = DIR_ENTRADA if idx == 0 else DIR_DESTINO
        is_entrada = (idx == 0)

        # Configura ponteiros locais para a thread usar (evita confusão se trocar de aba durante o scan)
        tree_alvo = self.tree_entrada if idx == 0 else self.tree_saida

        # Limpeza segura da lista antes de iniciar a thread
        if forcar_refresh:
            self.lbl_resumo.config(text="Lendo pasta...")

            # Limpa o visual imediatamente na thread principal
            for item in tree_alvo.get_children():
                tree_alvo.delete(item)

            # Limpa os dados em memória
            if idx == 0:
                self.dados_entrada = {}
                self.dados_atuais = self.dados_entrada
            else:
                self.dados_saida = {}
                self.dados_atuais = self.dados_saida

            # Inicia thread
            self.scan_ativo = True
            threading.Thread(target=self.thread_scan, args=(pasta_alvo, is_entrada, tree_alvo)).start()

    def registrar_log_excel(self, jobs_processados, resultados_impressao):
        """
        Gera/Atualiza o Excel com linhas repetidas para cada NF do CTE.
        """
        arquivo_excel = os.path.join(DIR_LOGS, "Relatorio_Impressoes.xlsx")

        cabecalhos = ["Data/Hora", "Usuário", "Chave CTE", "Chave NF", "Status", "Mensagem Erro"]

        # Tenta carregar ou cria novo
        if os.path.exists(arquivo_excel):
            try:
                wb = load_workbook(arquivo_excel)
                ws = wb.active

            except:
                wb = Workbook()
                ws = wb.active
                ws.append(cabecalhos)
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(cabecalhos)

        usuario = os.getlogin()
        agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        for job in jobs_processados:
            cte_key = job['cte_data']['chave'] if job['cte_data'] else "SEM_CTE"

            # Busca resultado
            res = resultados_impressao.get(cte_key, {'status': 'DESCONHECIDO', 'msg': ''})
            status_txt = res['status']
            msg_erro = res['msg']

            # Se for um job completo (CTE + NFs), criamos uma linha para cada NF
            if job['nfs_list']:
                for nf in job['nfs_list']:
                    nf_key = nf['chave']
                    ws.append([agora, usuario, cte_key, nf_key, status_txt, msg_erro])
            else:
                # Caso raro (ex: CTE avulso)
                ws.append([agora, usuario, cte_key, "---", status_txt, msg_erro])

        try:
            wb.save(arquivo_excel)
        except Exception as e:
            print(f"Erro ao salvar Excel: {e}")

    def thread_scan(self, pasta, is_entrada, tree_alvo):
        try:
            if is_entrada:
                self.update_status("Organizando arquivos...")
                self.separar_paginas_agrupadas(pasta)

            self.update_status("Listando arquivos...")
            arquivos = [f for f in os.listdir(pasta) if f.lower().endswith('.pdf')]

            # Dicionários para controle
            pool_nfs = {}  # Chave -> Dados do Original
            pool_ctes = []  # Lista de CTEs

            # Lock para garantir segurança na exclusão e verificação
            lock_dados = threading.Lock()

            total = len(arquivos)
            processados = 0
            duplicadas_excluidas = 0

            self.update_status(f"Analisando {total} arquivos...")
            max_workers = min(32, os.cpu_count() + 4)

            # --- ETAPA 1: Análise Paralela ---
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_file = {
                    executor.submit(self.analisar_pdf, os.path.join(pasta, f)): f
                    for f in arquivos
                }

                for future in concurrent.futures.as_completed(future_to_file):
                    f = future_to_file[future]
                    caminho_completo = os.path.join(pasta, f)
                    try:
                        tipo, chave_doc, referencias, pgs_reais, cliente_nome = future.result()
                        pgs = pgs_reais

                        with lock_dados:
                            if tipo == "NF" and chave_doc:
                                # VERIFICAÇÃO DE DUPLICIDADE
                                if chave_doc in pool_nfs:
                                    # JÁ EXISTE! EXCLUIR AUTOMATICAMENTE
                                    try:
                                        os.remove(caminho_completo)
                                        duplicadas_excluidas += 1
                                        continue
                                    except Exception:
                                        pass
                                else:
                                    # É o Original
                                    pool_nfs[chave_doc] = {'path': f, 'pgs': pgs, 'matched': False, 'chave': chave_doc, 'cliente': cliente_nome}

                            elif tipo == "CTE":
                                chave_cte = chave_doc if chave_doc else f"CTE_SEM_ID_{f}"
                                pool_ctes.append({'chave': chave_cte, 'path': f, 'pgs': pgs, 'refs': referencias})

                            elif not is_entrada and tipo == "DESCONHECIDO":
                                pool_nfs[f"GENERIC_{f}"] = {'path': f, 'pgs': pgs, 'matched': False,
                                                            'chave': f"Arquivo: {f}, 'cliente': False"}

                    except Exception as exc:
                        print(f'{f} gerou exceção: {exc}')

                    processados += 1
                    if processados % 10 == 0:
                        self.update_status(f"Analisando {processados}/{total}...")

            if duplicadas_excluidas > 0:
                print(f"Total de duplicatas removidas nesta varredura: {duplicadas_excluidas}")

            # --- ETAPA 2: Cruzamento (Montar Originais) ---
            self.update_status("Cruzando dados...")

            lista_originais = []

            # 2.1 Processar CTEs
            for cte in pool_ctes:
                nfs_vinculadas = []
                total_pgs_job = cte['pgs']

                for ref_key in cte['refs']:
                    if ref_key in pool_nfs:
                        nf_data = pool_nfs[ref_key]
                        if nf_data not in nfs_vinculadas:
                            nfs_vinculadas.append(nf_data)
                            nf_data['matched'] = True
                            total_pgs_job += nf_data['pgs']

                qtd = len(nfs_vinculadas)
                eh_completo = (qtd > 0)
                prioridade = 0 if eh_completo else 2

                lista_originais.append({
                    'prioridade': prioridade,
                    'tipo_obj': 'CTE_BASE',
                    'cte_data': cte,
                    'nfs_list': nfs_vinculadas,
                    'total_pgs': total_pgs_job,
                    'eh_completo': eh_completo,
                    'sort_name': cte['path']
                })

            # 2.2 Processar NFs Avulsas
            for chave_nf, dados in pool_nfs.items():
                if not dados['matched']:
                    lista_originais.append({
                        'prioridade': 1,
                        'tipo_obj': 'NF_AVULSA',
                        'nf_data': dados,
                        'total_pgs': dados['pgs'],
                        'eh_completo': False,
                        'sort_name': dados['path']
                    })

            # Ordena
            lista_originais.sort(key=lambda x: (x['prioridade'], x['sort_name']))

            # --- ETAPA 3: Inserção na Tabela ---
            novos_dados = {}
            id_counter = 1

            for item_orig in lista_originais:
                item_id = str(id_counter)
                dados_memoria = self._preparar_dados_visualizacao(item_orig, pasta, is_entrada)
                self._inserir_na_tree(tree_alvo, item_id, dados_memoria, id_counter, is_entrada)
                novos_dados[item_id] = dados_memoria
                id_counter += 1

            if is_entrada:
                self.dados_entrada.update(novos_dados)
            else:
                self.dados_saida.update(novos_dados)

            self.update_status("Pronto.")
            self.root.after(100, self.calcular_resumo)

        except Exception as e:
            print(f"Erro no Scan: {e}")
            import traceback
            traceback.print_exc()
            self.update_status("Erro ao ler arquivos.")
        finally:
            self.scan_ativo = False

    def _preparar_dados_visualizacao(self, item, pasta, is_entrada):
        dados = {}

        # Define visual do Cliente (Padrão)
        nome_exibicao = CLIENTE_PADRAO

        if item['tipo_obj'] == 'CTE_BASE':
            # Tenta pegar o nome do cliente da primeira NF vinculada
            if item['nfs_list']:
                nome_exibicao = item['nfs_list'][0].get('cliente', CLIENTE_PADRAO)

            cte = item['cte_data']
            nfs = item['nfs_list']
            num_cte = self._extrair_numero_documento(cte['chave'])

            # 1. Cria o dicionário base
            dados = {
                'tipo': 'COMPLETO' if item['eh_completo'] else 'CTE_AVULSO',
                'cte_data': cte,
                'nfs_list': nfs,
                'total_pgs': item['total_pgs'],
                'folder': pasta,
                'completo': item['eh_completo'],
                'search_text': f"{cte['chave']} {cte['path']} {num_cte} {nome_exibicao}".upper(),

                'col_visual_cli': nome_exibicao,
                'col_visual_cte': f"{cte['path']} ({cte['pgs']}pg)"
            }

            # 2. Preenche as colunas visuais dependendo se tem NF vinculada ou não
            if len(nfs) > 0:
                dados['col_visual_chave'] = "\n".join([nf['chave'] for nf in nfs])
                dados['col_visual_num'] = "\n".join([self._extrair_numero_documento(nf['chave']) for nf in nfs])
                dados['col_visual_nf'] = f"{nfs[0]['path']} ({nfs[0]['pgs']}pg)" if len(
                    nfs) == 1 else f"{len(nfs)} NFs vinculadas"

                # Adiciona dados das NFs na busca
                for nf in nfs:
                    dados['search_text'] += f" {nf['chave']} {nf['path']}".upper()
            else:
                dados['col_visual_chave'] = "---"
                dados['col_visual_num'] = num_cte + " (CTE)"
                dados['col_visual_nf'] = "--- (Faltando NF)"

            # 3. Define tags
            if is_entrada:
                dados['tag_visual'] = "checked" if item['eh_completo'] else "disabled"
            else:
                dados['tag_visual'] = "unchecked" if item['eh_completo'] else "disabled"

        elif item['tipo_obj'] == 'NF_AVULSA':
            nf = item['nf_data']
            nome_exibicao = nf.get('cliente', CLIENTE_PADRAO)

            num_nf = self._extrair_numero_documento(nf['chave'])

            dados = {
                'tipo': 'NF_AVULSA',
                'cte_data': None,
                'nfs_list': [nf],
                'total_pgs': item['total_pgs'],
                'folder': pasta,
                'completo': False,
                'search_text': f"{nf['chave']} {nf['path']} {num_nf} {nome_exibicao}".upper(),

                'col_visual_cli': nome_exibicao,
                'col_visual_num': num_nf,
                'col_visual_chave': nf['chave'],
                'col_visual_nf': f"{nf['path']} ({nf['pgs']}pg)",
                'col_visual_cte': "---",
                'tag_visual': "disabled" if is_entrada else "unchecked"
            }

        return dados

    def _inserir_na_tree(self, tree, iid, dados, counter, is_entrada):
        """Insere na Treeview aplicando as tags corretas"""
        valores = (
            iid,
            dados['col_visual_cli'],  # <--- COLUNA NOVA
            dados['col_visual_num'],
            dados['col_visual_chave'],
            dados['col_visual_nf'],
            dados['col_visual_cte']
        )

        tag_principal = dados['tag_visual']
        bg_tag = "odd_row" if (counter % 2 != 0) else "even_row"

        # Opcional: Adicionar tag se for Schneider para mudar cor da fonte
        tags_finais = (tag_principal, bg_tag)
        if dados.get('is_schneider'):
            tags_finais = (tag_principal, bg_tag, "schneider_row")

        tree.insert("", "end", iid=iid, tags=tags_finais, values=valores)


    def filtrar_tabela(self, *args):
        termo_bruto = self.var_busca.get().strip()
        termo_numerico = re.sub(r'[^0-9]', '', termo_bruto)

        # Se estiver vazio, exibe tudo
        if not termo_numerico:
            # Caso especial: se o usuario digitou letras, mas limpou os numeros, e o campo nao esta vazio
            # (ex: digitou "abc"), consideramos vazio para não travar a lista
            if termo_bruto and not termo_numerico:
                pass  # Continua filtrando (vai dar 0 resultados provavelmente)
            else:
                termo_numerico = ""  # Garante vazio

        # Limpa visualmente a tabela atual
        for item in self.tree_atual.get_children():
            self.tree_atual.delete(item)

        dados = self.dados_atuais
        if not dados: return

        ids_ordenados = sorted(dados.keys(), key=lambda x: int(x))

        for item_id in ids_ordenados:
            job = dados[item_id]

            match_encontrado = False

            if not termo_numerico:
                match_encontrado = True
            else:
                numeros_visuais = job.get('col_visual_num', '')

                # Limpa os pontos/quebras de linha do alvo para comparar apenas números
                # Ex alvo: "001234005678"
                alvo_limpo = re.sub(r'[^0-9]', '', numeros_visuais)

                # Verifica se o numero digitado está contido no alvo
                if termo_numerico in alvo_limpo:
                    match_encontrado = True

            # Se encontrou, insere na tabela
            if match_encontrado:
                # ... (recupera outros campos) ...
                col_acao = job.get('col_visual_acao', '') # <-- IMPORTANTE

                # Recalcula tags
                if job.get('tipo') == 'DUPLICATA':
                    tags_finais = ("duplicate_row",)
                else:
                    # Logica padrao checked/disabled
                    tag_status = job['tag_visual'] # Pegamos do cache pra facilitar
                    bg_tag = "odd_row" if (int(item_id) % 2 != 0) else "even_row"
                    tags_finais = (tag_status, bg_tag)

                self.tree_atual.insert("", "end", iid=item_id, tags=tags_finais,
                                       values=(item_id,
                                               job['col_visual_cli'],
                                               job['col_visual_num'],
                                               job['col_visual_chave'],
                                               job['col_visual_nf'],
                                               job['col_visual_cte']))

        self.calcular_resumo()

    def separar_paginas_agrupadas(self, pasta):
        arquivos = [x for x in os.listdir(pasta) if x.lower().endswith('.pdf')]

        for f in arquivos:
            full_path = os.path.join(pasta, f)

            try:
                reader = PdfReader(full_path)
                num_pages = len(reader.pages)

                # Se arquivo estiver vazio ou corrompido, ignora
                if num_pages == 0:
                    continue

                paginas_por_chave = {}
                chave_atual = "DESCONHECIDO"

                # Analisa as páginas
                for page in reader.pages:
                    txt = (page.extract_text() or "").upper()
                    chaves_na_pagina = self._encontrar_chaves_validas(txt)

                    if chaves_na_pagina:
                        chave_atual = chaves_na_pagina[0]

                    if chave_atual not in paginas_por_chave:
                        paginas_por_chave[chave_atual] = []
                    paginas_por_chave[chave_atual].append(page)

                qtd_chaves = len(paginas_por_chave)

                # CENÁRIO 1: PDF contém vários documentos (Split necessário)
                if qtd_chaves > 1:
                    for chave, paginas in paginas_por_chave.items():
                        writer = PdfWriter()
                        for p in paginas:
                            writer.add_page(p)

                        nome_final = f"{chave}.pdf" if chave != "DESCONHECIDO" else f"Split_{int(time.time())}_{f}"

                        # Salva o novo arquivo separado
                        with open(os.path.join(pasta, nome_final), "wb") as out:
                            writer.write(out)

                    # Fecha o arquivo original e DELETA ele (já que foi separado)
                    reader.stream.close()
                    try:
                        os.remove(full_path)
                    except Exception as e:
                        print(f"Não foi possível apagar o original {f}: {e}")

                # CENÁRIO 2: PDF contém apenas 1 documento (Apenas renomear)
                elif qtd_chaves == 1:
                    chave_encontrada = list(paginas_por_chave.keys())[0]

                    if chave_encontrada != "DESCONHECIDO":
                        nome_ideal = f"{chave_encontrada}.pdf"

                        # Se o nome já não for o ideal, renomeia
                        if f != nome_ideal:
                            reader.stream.close()  # Libera o arquivo
                            destino_ideal = os.path.join(pasta, nome_ideal)

                            if os.path.exists(destino_ideal):
                                # Se já existe um arquivo com essa chave, apaga o atual (duplicado)
                                os.remove(full_path)
                            else:
                                try:
                                    os.rename(full_path, destino_ideal)
                                except:
                                    pass

            except Exception as e:
                print(f"Erro ao separar {f}: {e}")

    @staticmethod
    def _encontrar_chaves_validas(texto_bruto):
        """
        Busca TODAS as chaves de acesso no documento (acumulativa),
        sem parar na primeira encontrada.
        """
        chaves_encontradas = []

        # --- ESTRATÉGIA 1: Padrão Visual (blocos de 4 dígitos com espaço ou ponto) ---
        # Geralmente captura a chave do CTE no cabeçalho
        padrao_visual = re.findall(r'(?:\d{4}[\s\.]){10}\d{4}', texto_bruto)

        for candidato in padrao_visual:
            limpo = re.sub(r'[^0-9]', '', candidato)
            if len(limpo) == 44:
                if ImpressorDHL._validar_dv_modulo11(limpo):
                    if limpo not in chaves_encontradas:
                        chaves_encontradas.append(limpo)

        # --- ESTRATÉGIA 2: Proximidade da Label "CHAVE DE ACESSO" ---
        idx_label = texto_bruto.find("CHAVE DE ACESSO")
        if idx_label != -1:
            fragmento = texto_bruto[idx_label:idx_label + 120]
            digitos_frag = re.sub(r'[^0-9]', '', fragmento)
            for i in range(len(digitos_frag) - 43):
                candidata = digitos_frag[i: i + 44]
                if ImpressorDHL._validar_dv_modulo11(candidata):
                    if candidata not in chaves_encontradas:
                        chaves_encontradas.append(candidata)

        # --- ESTRATÉGIA 3: Sequência Contínua (Sem espaços) ---
        # Essencial para pegar a chave da NF no campo "DOCUMENTOS ORIGINÁRIOS"
        padrao_continuo = re.findall(r'\b\d{44}\b', texto_bruto)
        for candidato in padrao_continuo:
            if ImpressorDHL._validar_dv_modulo11(candidato):
                if candidato not in chaves_encontradas:
                    chaves_encontradas.append(candidato)

        return chaves_encontradas

    @staticmethod
    def _validar_dv_modulo11(chave):
        """Função auxiliar isolada para validar o DV da chave"""
        try:
            if len(chave) != 44: return False
            uf = int(chave[:2])
            if uf < 11 or uf > 53: return False  # Validação de UF

            chave_base = chave[:43]
            dv_original = int(chave[43])
            soma = 0
            peso = 2
            for d in reversed(chave_base):
                soma += int(d) * peso
                peso += 1
                if peso > 9: peso = 2
            resto = soma % 11
            dv_calculado = 0 if resto < 2 else 11 - resto
            return dv_calculado == dv_original
        except:
            return False

    @staticmethod
    def _extrair_numero_documento(chave):
        """
        Extrai o número da NF/CT-e (9 dígitos) a partir da chave de acesso (posições 25 a 34).
        Aplica formatação visual para facilitar a leitura.
        """
        try:
            if len(chave) == 44:
                # Pega os 9 dígitos brutos (ex: "001779954")
                numero_raw = chave[25:34]

                if len(numero_raw) == 9:
                    return f"{numero_raw[:3]}{numero_raw[3:6]}{numero_raw[6:]}"

                return numero_raw
            return "?"
        except:
            return "?"

    def analisar_pdf(self, caminho):
        try:
            reader = PdfReader(caminho)
            num_paginas = len(reader.pages)
            texto_completo = ""

            # Leitura das páginas
            for page in reader.pages:
                try:
                    texto_completo += (page.extract_text() or "") + "\n"
                except:
                    pass

            texto_upper = texto_completo.upper()

            # --- NOVA LÓGICA DINÂMICA DE IDENTIFICAÇÃO ---
            cliente_detectado = CLIENTE_PADRAO

            # Percorre o dicionário de configurações
            for nome_cli, config in CLIENTES_CONFIG.items():
                # Verifica se alguma das palavras-chave está no texto
                for keyword in config["keywords"]:
                    if keyword in texto_upper:
                        cliente_detectado = nome_cli
                        break  # Encontrou a palavra, para de procurar palavras

                if cliente_detectado != CLIENTE_PADRAO:
                    break  # Encontrou o cliente, para de procurar clientes
            # ---------------------------------------------

            chaves = self._encontrar_chaves_validas(texto_upper)

            # Tentativa com Rotação (se não achou chaves)
            if not chaves:
                texto_completo = ""
                for page in reader.pages:
                    try:
                        page.rotate(90)
                        texto_completo += (page.extract_text() or "") + "\n"
                    except:
                        pass
                texto_upper = texto_completo.upper()
                chaves = self._encontrar_chaves_validas(texto_upper)

                # Re-verifica cliente no texto rotacionado se ainda for genérico
                if cliente_detectado == CLIENTE_PADRAO:
                    for nome_cli, config in CLIENTES_CONFIG.items():
                        for keyword in config["keywords"]:
                            if keyword in texto_upper:
                                cliente_detectado = nome_cli
                                break
                        if cliente_detectado != CLIENTE_PADRAO:
                            break

            if chaves:
                t, c, r = self._processar_resultado(texto_upper, chaves)
                # Retorna o NOME do cliente (string) em vez de True/False
                return t, c, r, num_paginas, cliente_detectado

            return "ERRO", None, [], 0, CLIENTE_PADRAO

        except Exception as e:
            return "ERRO", None, [], 0, CLIENTE_PADRAO

    def _processar_resultado(self, texto_upper, todas_chaves):
        """Processa o texto final para definir se é NF ou CTE"""
        # Limpa espaços para achar "D A C T E"
        texto_limpo = texto_upper.replace(" ", "").replace("\n", "").replace(".", "")

        tipo = "DESCONHECIDO"
        chave_principal = None
        referencias = []

        if todas_chaves:
            chave_principal = todas_chaves[0]

            # Lista de termos expandida
            termos_cte_limpos = [
                "DACTE",
                "CONHECIMENTODETRANSPORTE",
                "CT-E",
                "TOMADORDOSERVICO",
                "TOMADORDOSERVIÇO",
                "MODALRODOVIARIO",
                "DECLAROQUERECEBI"  # Comum no canhoto do CTE
            ]

            if any(termo in texto_limpo for termo in termos_cte_limpos):
                tipo = "CTE"
                if len(todas_chaves) > 1:
                    # Remove a chave principal da lista de referências
                    referencias = [c for c in todas_chaves if c != chave_principal]
            else:
                tipo = "NF"
        else:
            if "DANFE" in texto_limpo:
                tipo = "NF_SEM_CHAVE_LEGIVEL"
            elif "DACTE" in texto_limpo:
                tipo = "CTE_SEM_CHAVE_LEGIVEL"

        return tipo, chave_principal, referencias

    # --- IMPRESSÃO / REIMPRESSÃO ---
    def executar_acao_principal(self):
        printer = self.combo_printer.get()
        if not printer: return messagebox.showwarning("Aviso", "Selecione uma impressora.")
        self._salvar_preferencias()

        jobs_a_imprimir = []
        for item_id in self.tree_atual.get_children():
            if "checked" in self.tree_atual.item(item_id, "tags"):
                # Busca no dicionário atualmente ativo (entrada ou saida)
                if item_id in self.dados_atuais:
                    jobs_a_imprimir.append(self.dados_atuais[item_id])

        if not jobs_a_imprimir: return

        idx = self.notebook.index("current")
        is_reprint = (idx == 1)

        verbo = "REIMPRIMIR" if is_reprint else "IMPRIMIR"
        if messagebox.askyesno("Confirmar", f"{verbo} {len(jobs_a_imprimir)} conjuntos?"):
            threading.Thread(target=self.thread_print, args=(jobs_a_imprimir, printer, is_reprint)).start()

    def _verificar_se_schneider(self, caminho_arq):
        """
        Lê as primeiras páginas do PDF para verificar se a palavra SCHNEIDER
        está presente no conteúdo (Remetente, Destinatário, etc).
        """
        try:
            reader = PdfReader(caminho_arq)
            # Verifica apenas as 2 primeiras páginas para ganhar tempo
            paginas_para_ler = min(2, len(reader.pages))

            for i in range(paginas_para_ler):
                texto = (reader.pages[i].extract_text() or "").upper()
                if "SCHNEIDER" in texto:
                    return True
            return False
        except Exception as e:
            print(f"Erro ao verificar cliente no arquivo {os.path.basename(caminho_arq)}: {e}")
            return False

    def _gerar_temp_com_recusa(self, caminho_nf, caminho_recusa):
        """
        Gera PDF temporário intercalando página da NF + página de Recusa.
        """
        try:
            reader_nf = PdfReader(caminho_nf)
            reader_recusa = PdfReader(caminho_recusa)

            if len(reader_recusa.pages) == 0:
                return None

            pg_recusa = reader_recusa.pages[0]
            writer = PdfWriter()

            for page in reader_nf.pages:
                writer.add_page(page)  # Frente: Nota Fiscal
                writer.add_page(pg_recusa)  # Verso: Formulário de Recusa

            timestamp = int(time.time() * 100000)
            nome_temp = f"Temp_Print_{timestamp}.pdf"
            caminho_temp = os.path.join(DIR_LOGS, nome_temp)

            with open(caminho_temp, "wb") as f_out:
                writer.write(f_out)

            return caminho_temp
        except Exception as e:
            print(f"Erro ao gerar verso com recusa: {e}")
            return None

    def thread_print(self, lista_jobs, printer, is_reprint):
        if not os.path.exists(SUMATRA_PATH):
            messagebox.showerror("Erro", "SumatraPDF não encontrado.")
            return

        path_recusa = os.path.join(BASE_DIR, "Formulario_Recusa.pdf")
        existe_recusa = os.path.exists(path_recusa)

        cmd_simplex = f'"{SUMATRA_PATH}" -print-to "{printer}" -print-settings "simplex" -silent -exit-on-print'
        cmd_duplex = f'"{SUMATRA_PATH}" -print-to "{printer}" -print-settings "duplex" -silent -exit-on-print'

        resultados_log = {}
        total = len(lista_jobs)

        for i, job in enumerate(lista_jobs, 1):
            chave_log = job['cte_data']['chave'] if job['cte_data'] else f"Job_{i}"
            pasta_origem = job['folder']
            self.update_status(f"Imprimindo {i}/{total}: {chave_log}...")

            arquivos_movimentar = []
            erro_job = None

            # 1. CTE (SEMPRE SIMPLEX)
            if job['cte_data']:
                cte_path = os.path.join(pasta_origem, job['cte_data']['path'])
                if os.path.exists(cte_path):
                    try:
                        subprocess.run(f'{cmd_simplex} "{cte_path}"', shell=True, check=True)
                        arquivos_movimentar.append(cte_path)
                        time.sleep(1)
                    except Exception as e:
                        erro_job = f"Erro CTE: {e}"

            # 2. NFs (LÓGICA BASEADA NO NOME DO CLIENTE)
            for nf in job['nfs_list']:
                full_path = os.path.join(pasta_origem, nf['path'])
                if os.path.exists(full_path):
                    try:
                        # Recupera o nome do cliente salvo no scan
                        nome_cliente = nf.get('cliente', CLIENTE_PADRAO)

                        # Consulta as regras no dicionário GLOBAL
                        # Se não achar o cliente, usa False como padrão para recusa
                        config_cli = CLIENTES_CONFIG.get(nome_cliente, {"recusa": False})
                        precisa_recusa = config_cli["recusa"]

                        arquivo_final = full_path
                        comando_final = cmd_simplex
                        temp_criado = None

                        # SE O CLIENTE EXIGIR RECUSA E TIVERMOS O ARQUIVO
                        if precisa_recusa and existe_recusa:
                            temp = self._gerar_temp_com_recusa(full_path, path_recusa)
                            if temp:
                                arquivo_final = temp
                                comando_final = cmd_duplex
                                temp_criado = temp
                        elif precisa_recusa and not existe_recusa:
                            print(f"AVISO: Cliente {nome_cliente} exige recusa, mas arquivo não encontrado.")

                        subprocess.run(f'{comando_final} "{arquivo_final}"', shell=True, check=True)
                        arquivos_movimentar.append(full_path)

                        if temp_criado:
                            try:
                                time.sleep(0.5)
                                os.remove(temp_criado)
                            except:
                                pass

                    except Exception as e:
                        if not erro_job: erro_job = f"Erro NF: {e}"
                        print(f"Erro NF: {e}")

            # ... (Resto da lógica de Log e Movimentação permanece igual) ...
            if erro_job:
                resultados_log[chave_log] = {'status': 'ERRO', 'msg': str(erro_job)}
            else:
                resultados_log[chave_log] = {'status': 'SUCESSO', 'msg': ''}

            if not is_reprint and not erro_job and arquivos_movimentar:
                time.sleep(1)
                agora_ts = int(time.time())
                for origem in arquivos_movimentar:
                    try:
                        nome_original = os.path.basename(origem)
                        novo_nome = f"{agora_ts}___{nome_original}"
                        shutil.move(origem, os.path.join(DIR_DESTINO, novo_nome))
                    except:
                        pass

        self.update_status("Finalizado.")
        try:
            self.registrar_log_excel(lista_jobs, resultados_log)
        except:
            pass
        messagebox.showinfo("Concluído", "Impressão finalizada.")
        self.root.after(0, lambda: self.iniciar_scan(forcar_refresh=True))

    def _gerar_temp_com_recusa(self, caminho_nf, caminho_recusa):
        """
        Cria um PDF temporário onde:
        Página 1 = NF Pág 1
        Página 2 = Recusa
        Página 3 = NF Pág 2
        Página 4 = Recusa
        ...
        """
        try:
            reader_nf = PdfReader(caminho_nf)
            reader_recusa = PdfReader(caminho_recusa)

            # Pega a primeira página do formulário de recusa
            if len(reader_recusa.pages) > 0:
                pg_recusa = reader_recusa.pages[0]
            else:
                return None  # Arquivo de recusa vazio

            writer = PdfWriter()

            for page in reader_nf.pages:
                writer.add_page(page)  # Frente: Nota Fiscal
                writer.add_page(pg_recusa)  # Verso: Formulário de Recusa

            # Gera nome único para o temporário
            timestamp = int(time.time() * 1000)
            nome_temp = f"Temp_Print_{timestamp}.pdf"
            caminho_temp = os.path.join(DIR_LOGS, nome_temp)  # Salva na pasta de Logs temporariamente

            with open(caminho_temp, "wb") as f_out:
                writer.write(f_out)

            return caminho_temp

        except Exception as e:
            print(f"Erro ao gerar verso com recusa: {e}")
            return None

    def abrir_janela_ajuda(self):
        """Abre janela de ajuda Modal, Fixa e com Rolagem Estável"""
        help_win = tk.Toplevel(self.root)
        help_win.title("Ajuda e Instruções")
        help_win.configure(bg=BG_WHITE)

        # Configurações iniciais
        help_win.resizable(False, False)
        help_win.transient(self.root)  # Mantém a janela sempre na frente da principal

        try:
            help_win.iconbitmap(resource_path("logo.ico"))
        except:
            pass

        # Centralizar
        largura, altura = 700, 600
        pos_x = (help_win.winfo_screenwidth() // 2) - (largura // 2)
        pos_y = (help_win.winfo_screenheight() // 2) - (altura // 2)
        help_win.geometry(f"{largura}x{altura}+{pos_x}+{pos_y}")

        # --- CABEÇALHO ---
        header = tk.Frame(help_win, bg=DHL_YELLOW, height=60)
        header.pack(fill=tk.X, side=tk.TOP)

        title_frame = tk.Frame(header, bg=DHL_YELLOW)
        title_frame.pack(expand=True, fill=tk.BOTH)

        tk.Label(title_frame, text="COMO USAR O SISTEMA", bg=DHL_YELLOW, fg=DHL_RED,
                 font=("Segoe UI", 14, "bold")).pack(expand=True, pady=15)

        # --- ÁREA DE CONTEÚDO (CANVAS) ---
        container = tk.Frame(help_win, bg=BG_WHITE)
        container.pack(fill=tk.BOTH, expand=True)

        scrollbar = AutoScrollbar(container, orient="vertical")
        canvas = tk.Canvas(container, bg=BG_WHITE, highlightthickness=0, yscrollcommand=scrollbar.set)
        scrollbar.config(command=canvas.yview)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        inner_frame = tk.Frame(canvas, bg=BG_WHITE)
        canvas_window = canvas.create_window((0, 0), window=inner_frame, anchor="nw")

        def _on_mousewheel(event):
            try:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except:
                pass

        # Vincula scroll a tudo
        help_win.bind("<MouseWheel>", _on_mousewheel)

        def configure_frame(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        inner_frame.bind("<Configure>", configure_frame)
        canvas.itemconfigure(canvas_window, width=largura - 25)

        # --- CONTEÚDO ---
        content_box = tk.Frame(inner_frame, bg=BG_WHITE, padx=20, pady=20)
        content_box.pack(fill=tk.BOTH, expand=True)

        steps = [
            ("1. PREPARAÇÃO",
             "Mova os arquivos PDF (NFs e CTEs) para a pasta '1_PastaOrigem'."),
            ("2. CARREGAMENTO",
             "Clique no botão 'ATUALIZAR LISTA' e aguarde o programa ler e organizar os arquivos."),
            ("3. CONFERÊNCIA E FILTRO",
             "Verifique a lista na tela. Use o campo de busca (ícone da Lupa) para filtrar rapidamente por um número de Nota Fiscal específico."),
            ("4. IMPRESSÃO E PREFERÊNCIAS",
             "Selecione a impressora e clique em 'IMPRIMIR AGORA'. O sistema memoriza automaticamente sua última impressora escolhida."),
            ("5. AGUARDE A CONCLUSÃO",
             "Não feche o programa. Observe o progresso na faixa preta inferior e aguarde o aviso de confirmação. Os arquivos serão movidos automaticamente."),
            ("6. REIMPRESSÃO (CORREÇÃO)",
             "Se houver falha no papel ou atolamento, vá para a aba 'IMPRESSÕES REALIZADAS' para reimprimir apenas os documentos necessários."),
            ("7. HISTÓRICO AUTOMÁTICO",
             "Arquivos processados ficam na pasta de Destino por 24h. Após esse período, são movidos automaticamente para a pasta 'Histórico'."),
            ("8. ACESSOS E SUPORTE",
             "O acesso às pastas de rede é restrito. Para liberar novos usuários ou resolver erros de permissão, é obrigatório contatar o suporte técnico abaixo.")
        ]

        for t, d in steps:
            f = tk.Frame(content_box, bg=BG_WHITE, pady=8)
            f.pack(fill=tk.X, anchor="w")
            tk.Label(f, text=t, font=("Segoe UI", 11, "bold"), fg=DHL_RED, bg=BG_WHITE).pack(anchor="w")
            tk.Label(f, text=d, font=("Segoe UI", 10), fg="#333", bg=BG_WHITE, justify="left",
                     wraplength=largura - 80).pack(anchor="w")

        tk.Frame(content_box, bg="#E0E0E0", height=2).pack(fill=tk.X, pady=20)

        # --- CONTATOS ---
        tk.Label(content_box, text="SUPORTE TÉCNICO", font=("Segoe UI", 11, "bold"), bg=BG_WHITE).pack(anchor="w",
                                                                                                       pady=(0, 5))
        tk.Label(content_box, text="Para dúvidas, erros ou sugestões, entre em contato:", bg=BG_WHITE,
                 font=("Segoe UI", 10)).pack(anchor="w")

        links_frame = tk.Frame(content_box, bg=BG_WHITE, pady=10)
        links_frame.pack(anchor="w")

        def send_mail(email):
            webbrowser.open(f"mailto:{email}")

        def criar_link(email):
            lbl = tk.Label(links_frame, text=f"{email}", fg="#0055A5", bg=BG_WHITE,
                           cursor="hand2", font=("Segoe UI", 10, "bold", "underline"))
            lbl.pack(anchor="w", pady=3)
            lbl.bind("<Button-1>", lambda e: send_mail(email))

        criar_link("eduardojuan@dhl.com")
        criar_link("aurelio.ferreira@dhl.com")

        tk.Label(content_box, text="", bg=BG_WHITE).pack(pady=10)

        # 1. Garante que a janela foi desenhada antes de tentar travar
        help_win.wait_visibility()

        # 2. Trava todos os eventos para esta janela (ninguém mexe no fundo)
        help_win.grab_set()

        # 3. Força o foco para o popup
        help_win.focus_set()

        # 4. Espera a janela fechar
        self.root.wait_window(help_win)

    def abrir_janela_log(self):
        """Abre janela de log Modal e Fixa"""
        log_win = tk.Toplevel(self.root)
        log_win.title("Log de Auditoria - Hoje")
        log_win.configure(bg=BG_WHITE)

        # Configurações iniciis
        log_win.resizable(False, False)
        log_win.transient(self.root)

        try:
            log_win.iconbitmap(resource_path("logo.ico"))
        except:
            pass

        # Centralizar
        largura_win = 700
        altura_win = 500
        largura_screen = log_win.winfo_screenwidth()
        altura_screen = log_win.winfo_screenheight()
        pos_x = (largura_screen // 2) - (largura_win // 2)
        pos_y = (altura_screen // 2) - (altura_win // 2)
        log_win.geometry(f"{largura_win}x{altura_win}+{pos_x}+{pos_y}")

        # --- Cabeçalho ---
        header = tk.Frame(log_win, bg=DHL_YELLOW, height=50)
        header.pack(fill=tk.X, side=tk.TOP)

        frame_head = tk.Frame(header, bg=DHL_YELLOW)
        frame_head.pack(expand=True, fill=tk.BOTH)

        tk.Label(frame_head, text="REGISTROS DE ATIVIDADE (HOJE)", bg=DHL_YELLOW, fg=DHL_RED,
                 font=("Segoe UI", 12, "bold")).pack(expand=True, pady=10)

        # --- Área de Texto e Scrollbar ---
        frame_texto = tk.Frame(log_win, bg=BG_WHITE)
        frame_texto.pack(fill=tk.BOTH, expand=True)

        scrollbar = AutoScrollbar(frame_texto)

        # Texto com padding interno
        txt_area = tk.Text(frame_texto, font=("Consolas", 9), bg="#F9F9F9", relief="flat",
                           yscrollcommand=scrollbar.set, padx=15, pady=15)

        scrollbar.config(command=txt_area.yview)

        # Scrollbar colada na direita
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        txt_area.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        conteudo = AuditLogger.ler_logs_do_dia()
        txt_area.insert(tk.END, conteudo)
        txt_area.configure(state='disabled')

        # 1. Garante que a janela foi desenhada antes de tentar travar
        log_win.wait_visibility()

        # 2. Trava todos os eventos para esta janela
        log_win.grab_set()

        # 3. Força o foco
        log_win.focus_set()

        # 4. Espera fechar
        self.root.wait_window(log_win)

if __name__ == "__main__":
    root = tk.Tk()
    app = ImpressorDHL(root)
    root.mainloop()